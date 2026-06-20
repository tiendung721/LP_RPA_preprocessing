from __future__ import annotations

import sys
from datetime import date

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

import update_rpa_status as update_cli
from src.flows import FLOW_BAO_CO, FLOW_BAO_NO, FLOW_CHI_TIEN_MAT, FLOW_THU_TIEN_MAT
from src.models import ProcessedTransaction
from src.output_writer import write_excel
from src.rpa_input_status import (
    INPUT_MESSAGE_COLUMN,
    INPUT_STATUS_COLUMN,
    INPUT_STATUS_COLUMNS,
    INPUT_UPDATED_AT_COLUMN,
    RPA_INPUT_SHEETS,
    TRANSACTION_UID_COLUMN,
    RpaInputStatusError,
    update_input_file_status,
)
from src.rpa_summary import SUMMARY_COLUMNS, SUMMARY_SHEET_NAME, write_summary
from src.rpa_tracking import STATUS_DONE, STATUS_PENDING


def _processed(uid: str, flow: str, amount: int) -> ProcessedTransaction:
    return ProcessedTransaction(
        source_file="sample.xlsx",
        original_row_index=2,
        bank="ACB",
        flow=flow,
        transaction_date=date(2026, 6, 20),
        object_code="ABC",
        object_name="ABC",
        reason="Thanh toan ABC",
        debit_account="331",
        credit_account="1121CT",
        amount=amount,
        use_case="Thanh toan",
        original_content=f"TT ABC {uid}",
        counterparty_raw="ABC",
        doc_no=uid,
        status="OK",
        error_note="",
        confidence=0.95,
        transaction_uid=uid,
        source_sheet="Statement",
    )


def _make_input_workbook(path, uid: str = "uid_1", include_status: bool = True, target_sheet: str = "BAO_NO_INPUT") -> None:
    workbook = Workbook()
    for index, sheet_name in enumerate(RPA_INPUT_SHEETS):
        ws = workbook.active if index == 0 else workbook.create_sheet(sheet_name)
        ws.title = sheet_name
        headers = ["Lí do", TRANSACTION_UID_COLUMN, "Thành tiền"]
        values = ["Manual reason", uid if sheet_name == target_sheet else f"{uid}_{sheet_name}", 1000]
        if include_status:
            headers += INPUT_STATUS_COLUMNS
            values += [STATUS_PENDING, "old message", ""]
        ws.append(headers)
        ws.append(values)
    workbook.save(path)


def _sheet_row(path, sheet_name: str = "BAO_NO_INPUT") -> tuple[list[str], dict[str, object]]:
    workbook = load_workbook(path, data_only=True)
    ws = workbook[sheet_name]
    headers = [cell.value for cell in ws[1]]
    values = dict(zip(headers, [cell.value for cell in ws[2]]))
    return headers, values


def _write_summary(path, uid: str) -> None:
    row = {column: "" for column in SUMMARY_COLUMNS}
    row.update({"transaction_uid": uid, "status": STATUS_PENDING, "rpa_status": STATUS_PENDING})
    write_summary(pd.DataFrame([row], columns=SUMMARY_COLUMNS), path)


def test_generated_rpa_input_has_status_columns_for_all_input_sheets(tmp_path):
    output_file = tmp_path / "rpa_input.xlsx"
    items = [
        _processed("uid_bao_no", FLOW_BAO_NO, 100),
        _processed("uid_bao_co", FLOW_BAO_CO, 200),
        _processed("uid_chi", FLOW_CHI_TIEN_MAT, 300),
        _processed("uid_thu", FLOW_THU_TIEN_MAT, 400),
    ]

    write_excel(items, output_file, run_id="run1")

    workbook = load_workbook(output_file, data_only=True)
    for sheet_name in RPA_INPUT_SHEETS:
        headers = [cell.value for cell in workbook[sheet_name][1]]
        values = dict(zip(headers, [cell.value for cell in workbook[sheet_name][2]]))
        assert headers[-3:] == INPUT_STATUS_COLUMNS
        assert values[INPUT_STATUS_COLUMN] == STATUS_PENDING
        assert values[INPUT_MESSAGE_COLUMN] in ("", None)
        assert values[INPUT_UPDATED_AT_COLUMN] in ("", None)


def test_update_input_file_status_updates_matching_uid(tmp_path):
    input_file = tmp_path / "rpa_input.xlsx"
    _make_input_workbook(input_file, uid="uid_done")

    result = update_input_file_status(input_file, "uid_done", STATUS_DONE, message="Nhap thanh cong")

    assert result["updated_rows"] == [{"sheet": "BAO_NO_INPUT", "row": 2}]
    _, values = _sheet_row(input_file)
    assert values[INPUT_STATUS_COLUMN] == STATUS_DONE
    assert values[INPUT_MESSAGE_COLUMN] == "Nhap thanh cong"
    assert values[INPUT_UPDATED_AT_COLUMN]
    assert values["Lí do"] == "Manual reason"


def test_update_input_file_status_adds_missing_status_columns(tmp_path):
    input_file = tmp_path / "rpa_input.xlsx"
    _make_input_workbook(input_file, uid="uid_done", include_status=False)

    update_input_file_status(input_file, "uid_done", STATUS_DONE, message="OK")

    headers, values = _sheet_row(input_file)
    assert headers[-3:] == INPUT_STATUS_COLUMNS
    assert values[INPUT_STATUS_COLUMN] == STATUS_DONE
    assert values[INPUT_MESSAGE_COLUMN] == "OK"
    assert values[INPUT_UPDATED_AT_COLUMN]


def test_update_input_file_status_raises_when_uid_not_found_without_saving(tmp_path):
    input_file = tmp_path / "rpa_input.xlsx"
    _make_input_workbook(input_file, uid="uid_existing", include_status=False)

    with pytest.raises(RpaInputStatusError, match="transaction_uid not found in input file: uid_missing"):
        update_input_file_status(input_file, "uid_missing", STATUS_DONE)

    headers, _ = _sheet_row(input_file)
    assert INPUT_STATUS_COLUMN not in headers


def test_update_status_cli_updates_multiple_input_files(tmp_path, monkeypatch):
    summary_path = tmp_path / "rpa_summary.xlsx"
    _write_summary(summary_path, "uid_done")
    input_file_1 = tmp_path / "rpa_input.xlsx"
    input_file_2 = tmp_path / "rpa_input_history.xlsx"
    _make_input_workbook(input_file_1, uid="uid_done")
    _make_input_workbook(input_file_2, uid="uid_done")
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "update_rpa_status.py",
            "--output-dir",
            str(tmp_path),
            "--input-file",
            str(input_file_1),
            "--input-file",
            str(input_file_2),
            "--uid",
            "uid_done",
            "--status",
            STATUS_DONE,
            "--message",
            "Nhap thanh cong",
            "--run-id",
            "run_1",
        ],
    )

    assert update_cli.main() == 0

    for input_file in (input_file_1, input_file_2):
        _, values = _sheet_row(input_file)
        assert values[INPUT_STATUS_COLUMN] == STATUS_DONE
        assert values[INPUT_MESSAGE_COLUMN] == "Nhap thanh cong"
        assert values[INPUT_UPDATED_AT_COLUMN]

    summary_df = pd.read_excel(summary_path, sheet_name=SUMMARY_SHEET_NAME, dtype=object)
    summary_df = summary_df.where(pd.notna(summary_df), "")
    row = summary_df[summary_df["transaction_uid"] == "uid_done"].iloc[0]
    assert row["rpa_status"] == STATUS_DONE
    assert row["rpa_message"] == "Nhap thanh cong"


def test_update_status_cli_returns_1_when_one_input_file_cannot_find_uid(tmp_path, monkeypatch, capsys):
    summary_path = tmp_path / "rpa_summary.xlsx"
    _write_summary(summary_path, "uid_done")
    good_input = tmp_path / "rpa_input.xlsx"
    bad_input = tmp_path / "rpa_input_other.xlsx"
    _make_input_workbook(good_input, uid="uid_done")
    _make_input_workbook(bad_input, uid="uid_other")
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "update_rpa_status.py",
            "--output-dir",
            str(tmp_path),
            "--input-file",
            str(good_input),
            "--input-file",
            str(bad_input),
            "--uid",
            "uid_done",
            "--status",
            STATUS_DONE,
            "--message",
            "Nhap thanh cong",
        ],
    )

    assert update_cli.main() == 1
    output = capsys.readouterr().out
    assert str(bad_input) in output
    assert "transaction_uid not found in input file: uid_done" in output

    _, good_values = _sheet_row(good_input)
    _, bad_values = _sheet_row(bad_input)
    assert good_values[INPUT_STATUS_COLUMN] == STATUS_DONE
    assert bad_values[INPUT_STATUS_COLUMN] == STATUS_PENDING


def test_update_status_cli_returns_1_for_missing_input_file(tmp_path, monkeypatch):
    summary_path = tmp_path / "rpa_summary.xlsx"
    _write_summary(summary_path, "uid_done")
    missing_file = tmp_path / "missing.xlsx"
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "update_rpa_status.py",
            "--output-dir",
            str(tmp_path),
            "--input-file",
            str(missing_file),
            "--uid",
            "uid_done",
            "--status",
            STATUS_DONE,
        ],
    )

    assert update_cli.main() == 1
    summary_df = pd.read_excel(summary_path, sheet_name=SUMMARY_SHEET_NAME, dtype=object)
    row = summary_df[summary_df["transaction_uid"] == "uid_done"].iloc[0]
    assert row["rpa_status"] == STATUS_PENDING
