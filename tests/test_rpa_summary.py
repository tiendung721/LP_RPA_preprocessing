from __future__ import annotations

from datetime import date

import pandas as pd
from openpyxl import load_workbook

from src.models import ProcessedTransaction
from src.output_writer import RPA_BUSINESS_COLUMNS, write_outputs
from src.rpa_summary import (
    STATUS_DONE,
    STATUS_PENDING,
    SUMMARY_COLUMNS,
    SUMMARY_SHEET_NAME,
    abort_rpa_run,
    finalize_rpa_run,
    mark_rpa_done,
    mark_rpa_started,
    write_summary,
)


def _processed(uid: str, row_index: int, amount: int = 1000) -> ProcessedTransaction:
    return ProcessedTransaction(
        source_file="sample.xlsx",
        original_row_index=row_index,
        bank="ACB",
        flow="bao_no",
        transaction_date=date(2026, 4, 1),
        object_code="ABC",
        object_name="ABC",
        reason="Thanh toán ABC",
        debit_account="331",
        credit_account="1121CT",
        amount=amount,
        use_case="Chi phí thanh toán",
        original_content=f"TT CHO ABC {row_index}",
        counterparty_raw="ABC",
        doc_no=str(row_index),
        status="OK",
        error_note="",
        confidence=0.95,
        transaction_uid=uid,
        source_sheet="Statement",
    )


def test_write_outputs_filters_completed_summary_rows(tmp_path):
    summary_path = tmp_path / "rpa_summary.xlsx"
    existing_summary = pd.DataFrame(
        [
            {
                "transaction_uid": "uid_done",
                "source_file": "sample.xlsx",
                "source_sheet": "Statement",
                "source_row_index": 2,
                "bank": "ACB",
                "flow": "bao_no",
                "transaction_date": date(2026, 4, 1),
                "doc_no": "2",
                "original_content": "TT CHO ABC 2",
                "counterparty_raw": "ABC",
                "amount": 1000,
                "object_code": "ABC",
                "object_name": "ABC",
                "debit_account": "331",
                "credit_account": "1121CT",
                "reason": "Thanh toán ABC",
                "status": STATUS_DONE,
                "last_run_id": "old_run",
                "rpa_started_at": "2026-04-01T08:00:00",
                "rpa_finished_at": "2026-04-01T08:01:00",
                "rpa_message": "",
                "voucher_no": "BN001",
                "created_at": "2026-04-01T08:00:00",
                "updated_at": "2026-04-01T08:01:00",
            }
        ],
        columns=SUMMARY_COLUMNS,
    )
    write_summary(existing_summary, summary_path)

    config = {
        "output": {
            "excel_file": "rpa_input.xlsx",
            "tracking_file": "rpa_tracking.json",
            "summary_file": "rpa_summary.xlsx",
        }
    }
    write_outputs([_processed("uid_done", 2), _processed("uid_new", 3)], tmp_path, config)

    wb = load_workbook(tmp_path / "rpa_input.xlsx", data_only=True)
    assert wb["BAO_NO_INPUT"].max_row == 2
    headers = [cell.value for cell in wb["BAO_NO_INPUT"][1]]
    values = dict(zip(headers, [cell.value for cell in wb["BAO_NO_INPUT"][2]]))
    assert values["Thành tiền"] == 1000
    assert wb["RPA_TASKS"].max_row == 2
    task_headers = [cell.value for cell in wb["RPA_TASKS"][1]]
    task_values = dict(zip(task_headers, [cell.value for cell in wb["RPA_TASKS"][2]]))
    assert task_values["transaction_uid"] == "uid_new"
    assert task_values["source_row_index"] == 3

    summary_df = pd.read_excel(summary_path, sheet_name=SUMMARY_SHEET_NAME, dtype=object)
    statuses = dict(zip(summary_df["transaction_uid"], summary_df["status"]))
    assert statuses["uid_done"] == STATUS_DONE
    assert statuses["uid_new"] == STATUS_PENDING


def test_write_outputs_reexports_legacy_non_done_rows_as_pending(tmp_path):
    summary_path = tmp_path / "rpa_summary.xlsx"
    existing_summary = pd.DataFrame(
        [
            {
                "transaction_uid": "uid_in_progress",
                "source_file": "sample.xlsx",
                "source_sheet": "Statement",
                "source_row_index": 2,
                "bank": "ACB",
                "flow": "bao_no",
                "status": "dang_nhap",
                "created_at": "2026-04-01T08:00:00",
                "updated_at": "2026-04-01T08:00:00",
            },
            {
                "transaction_uid": "uid_error",
                "source_file": "sample.xlsx",
                "source_sheet": "Statement",
                "source_row_index": 3,
                "bank": "ACB",
                "flow": "bao_no",
                "status": "loi",
                "rpa_message": "PAD nhap loi",
                "created_at": "2026-04-01T08:00:00",
                "updated_at": "2026-04-01T08:01:00",
            },
        ],
        columns=SUMMARY_COLUMNS,
    )
    write_summary(existing_summary, summary_path)

    config = {
        "output": {
            "excel_file": "rpa_input.xlsx",
            "tracking_file": "rpa_tracking.json",
            "summary_file": "rpa_summary.xlsx",
        }
    }
    result = write_outputs([_processed("uid_in_progress", 2), _processed("uid_error", 3)], tmp_path, config)

    wb = load_workbook(result.excel_path, data_only=True)
    headers = [cell.value for cell in wb["BAO_NO_INPUT"][1]]
    assert headers == RPA_BUSINESS_COLUMNS
    task_headers = [cell.value for cell in wb["RPA_TASKS"][1]]
    rows = [dict(zip(task_headers, row)) for row in wb["RPA_TASKS"].iter_rows(min_row=2, values_only=True)]
    statuses = {row["transaction_uid"]: row["summary_status"] for row in rows}
    assert statuses == {
        "uid_in_progress": STATUS_PENDING,
        "uid_error": STATUS_PENDING,
    }
    assert result.stats["in_progress_count"] == 0
    assert result.stats["error_count"] == 0
    assert result.stats["bao_no_output_count"] == 2


def test_write_outputs_promotes_legacy_success_attempts_to_done(tmp_path):
    summary_path = tmp_path / "rpa_summary.xlsx"
    existing_summary = pd.DataFrame(
        [
            {
                "transaction_uid": "uid_success",
                "source_file": "sample.xlsx",
                "source_sheet": "Statement",
                "source_row_index": 2,
                "bank": "ACB",
                "flow": "bao_no",
                "status": STATUS_PENDING,
                "last_attempt_result": "success",
                "rpa_finished_at": "2026-04-01T08:01:00",
                "created_at": "2026-04-01T08:00:00",
                "updated_at": "2026-04-01T08:01:00",
            }
        ],
        columns=SUMMARY_COLUMNS,
    )
    write_summary(existing_summary, summary_path)

    config = {
        "output": {
            "excel_file": "rpa_input.xlsx",
            "tracking_file": "rpa_tracking.json",
            "summary_file": "rpa_summary.xlsx",
        }
    }
    result = write_outputs([_processed("uid_success", 2)], tmp_path, config)

    wb = load_workbook(result.excel_path, data_only=True)
    assert wb["BAO_NO_INPUT"].max_row == 1
    assert wb["RPA_TASKS"].max_row == 1

    summary_df = pd.read_excel(summary_path, sheet_name=SUMMARY_SHEET_NAME, dtype=object)
    row = summary_df[summary_df["transaction_uid"] == "uid_success"].iloc[0]
    assert row["status"] == STATUS_DONE
    assert row["rpa_status"] == STATUS_DONE


def test_rpa_status_helpers_mark_done_immediately_completed(tmp_path):
    summary_path = tmp_path / "rpa_summary.xlsx"
    row = {
        column: ""
        for column in SUMMARY_COLUMNS
    }
    row.update(
        {
            "transaction_uid": "uid_pending",
            "source_file": "sample.xlsx",
            "source_sheet": "Statement",
            "source_row_index": 5,
            "status": STATUS_PENDING,
        }
    )
    write_summary(pd.DataFrame([row], columns=SUMMARY_COLUMNS), summary_path)

    mark_rpa_started(summary_path, "uid_pending", "run_1")
    df = pd.read_excel(summary_path, sheet_name=SUMMARY_SHEET_NAME, dtype=object)
    df = df.where(pd.notna(df), "")
    started_row = df[df["transaction_uid"] == "uid_pending"].iloc[0]
    assert started_row["status"] == STATUS_PENDING
    assert started_row["rpa_started_at"]

    mark_rpa_done(summary_path, "uid_pending", "run_1", voucher_no="BN001")
    df = pd.read_excel(summary_path, sheet_name=SUMMARY_SHEET_NAME, dtype=object)
    df = df.where(pd.notna(df), "")
    done_row = df[df["transaction_uid"] == "uid_pending"].iloc[0]
    assert done_row["status"] == STATUS_DONE
    assert done_row["last_attempt_result"] == ""
    assert done_row["voucher_no"] == "BN001"
    assert done_row["rpa_finished_at"]
    assert done_row["completed_at"]

    finalize_rpa_run(summary_path, "run_1")
    df = pd.read_excel(summary_path, sheet_name=SUMMARY_SHEET_NAME, dtype=object)
    df = df.where(pd.notna(df), "")
    finalized_row = df[df["transaction_uid"] == "uid_pending"].iloc[0]
    assert finalized_row["status"] == STATUS_DONE
    assert finalized_row["voucher_no"] == "BN001"
    assert finalized_row["completed_at"]
    assert finalized_row["last_attempt_result"] == ""


def test_abort_run_resets_only_rows_touched_in_that_run(tmp_path):
    summary_path = tmp_path / "rpa_summary.xlsx"
    rows = []
    for uid, status in (("uid_attempted", STATUS_PENDING), ("uid_untouched", STATUS_PENDING), ("uid_done_old", STATUS_DONE)):
        row = {column: "" for column in SUMMARY_COLUMNS}
        row.update(
            {
                "transaction_uid": uid,
                "source_file": "sample.xlsx",
                "source_sheet": "Statement",
                "source_row_index": 5,
                "status": status,
                "rpa_status": status,
                "voucher_no": "OLD" if uid == "uid_done_old" else "",
                "completed_at": "2026-04-01T08:01:00" if uid == "uid_done_old" else "",
            }
        )
        rows.append(row)
    write_summary(pd.DataFrame(rows, columns=SUMMARY_COLUMNS), summary_path)

    mark_rpa_started(summary_path, "uid_attempted", "run_abort")
    from src.rpa_summary import mark_rpa_error

    mark_rpa_error(summary_path, "uid_attempted", "run_abort", "VACOM row error")
    mark_rpa_done(summary_path, "uid_done_old", "run_abort", voucher_no="OLD")
    abort_rpa_run(summary_path, "run_abort", message="PAD abort")

    df = pd.read_excel(summary_path, sheet_name=SUMMARY_SHEET_NAME, dtype=object)
    df = df.where(pd.notna(df), "")
    statuses = dict(zip(df["transaction_uid"], df["status"]))
    vouchers = dict(zip(df["transaction_uid"], df["voucher_no"]))
    messages = dict(zip(df["transaction_uid"], df["rpa_message"]))

    assert statuses["uid_attempted"] == STATUS_PENDING
    assert vouchers["uid_attempted"] == ""
    assert messages["uid_attempted"] == "PAD abort"
    assert statuses["uid_untouched"] == STATUS_PENDING
    assert statuses["uid_done_old"] == STATUS_DONE
    assert vouchers["uid_done_old"] == "OLD"
