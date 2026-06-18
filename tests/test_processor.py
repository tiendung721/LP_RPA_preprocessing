from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from src.config_loader import load_config, load_rules
from src.ml.object_ranker import ObjectRanker
from src.models import ExtractedEntities, ObjectCandidate, ProcessedTransaction, Transaction
from src.object_matcher import ObjectMatcher
from src.output_writer import RPA_BUSINESS_COLUMNS, RPA_COLUMNS, write_excel, write_object_match_review
from src.processor import process_transaction
from src.reason_generator import generate_reason, reason_requires_object_code
from src.rule_engine import RuleEngine


PROJECT_ROOT = Path(__file__).resolve().parents[1]


def _config():
    return load_config(PROJECT_ROOT / "config/config.yaml")


def _engine():
    rules, _ = load_rules(None, PROJECT_ROOT / "config/default_rules.yaml")
    return RuleEngine(rules)


def test_both_debit_credit_goes_exception():
    txn = Transaction(
        source_file="sample.xlsx",
        bank="ACB",
        transaction_date=date(2026, 4, 1),
        doc_no="1",
        description="THANH TOAN ABC",
        counterparty_raw="",
        debit_amount=100,
        credit_amount=100,
        original_row_index=2,
    )
    result = process_transaction(txn, _config(), _engine(), ObjectMatcher([]), ObjectMatcher([]))
    assert result.status == "ERROR"
    assert "Dòng có cả ghi nợ và ghi có" in result.error_note


class _FakeObjectModel:
    classes_ = [0, 1]

    def predict_proba(self, texts):
        return [[0.05, 0.95] if "RIGHT" in text else [0.9, 0.1] for text in texts]


def test_process_transaction_uses_object_ranker_for_ambiguous_candidates():
    txn = Transaction(
        source_file="sample.xlsx",
        bank="ACB",
        transaction_date=date(2026, 4, 1),
        doc_no="1",
        description="LE PHAM TT CHO TARGET",
        counterparty_raw="",
        debit_amount=1000,
        credit_amount=0,
        original_row_index=2,
    )
    payable = ObjectMatcher.from_records(
        [
            {"code": "WRONG", "name": "Công ty WRONG TARGET"},
            {"code": "RIGHT", "name": "Công ty RIGHT TARGET"},
        ],
        min_score=80,
        min_gap=8,
    )
    ranker = ObjectRanker(enabled=False, min_confidence=0.85, min_gap=0.15)
    ranker.model = _FakeObjectModel()

    result = process_transaction(txn, _config(), _engine(), ObjectMatcher([]), payable, object_ranker=ranker)

    assert result.status == "OK"
    assert result.object_code == "RIGHT"
    assert result.object_match_source == "ml_object_ranker"
    assert result.object_ml_result.status == "OK"


def test_generate_reason_uses_business_template_and_object_code():
    assert generate_reason("bao_no", "331", "1121CT", "PIL") == "Thanh toán công nợ PIL"
    assert generate_reason("bao_no", "141", "1121CT", "KHÁCH A") == "Tạm ứng cá nhân KHÁCH A"
    assert generate_reason("bao_no", "334", "1121CT", "") == "Trả lương nhân viên"
    assert generate_reason("bao_co", "1121VCB", "131", "KVIII") == "Thu tiền công nợ KVIII"
    assert generate_reason("bao_co", "1121VCB", "515", "") == "Lãi tiền gửi ngân hàng"
    assert reason_requires_object_code("bao_no", "331", "1121CT")
    assert not reason_requires_object_code("bao_no", "334", "1121CT")


def test_process_transaction_reason_does_not_use_object_name():
    txn = Transaction(
        source_file="sample.xlsx",
        bank="ACB",
        transaction_date=date(2026, 4, 1),
        doc_no="1",
        description="THANH TOAN KHU VUC III",
        counterparty_raw="Khu Vuc III",
        debit_amount=1000,
        credit_amount=0,
        original_row_index=2,
    )
    payable = ObjectMatcher.from_records(
        [
            {"code": "PIL", "name": "Cong ty Khu Vuc III"},
        ],
        min_score=80,
        min_gap=8,
    )

    result = process_transaction(txn, _config(), _engine(), ObjectMatcher([]), payable)

    assert result.status == "OK"
    assert result.object_code == "PIL"
    assert result.object_name == "Cong ty Khu Vuc III"
    assert result.reason == "Thanh toán công nợ PIL"


def test_missing_object_code_required_for_reason_goes_exception(tmp_path):
    txn = Transaction(
        source_file="sample.xlsx",
        bank="ACB",
        transaction_date=date(2026, 4, 1),
        doc_no="1",
        description="THANH TOAN KHU VUC III",
        counterparty_raw="Khu Vuc III",
        debit_amount=1000,
        credit_amount=0,
        original_row_index=2,
    )

    result = process_transaction(txn, _config(), _engine(), ObjectMatcher([]), ObjectMatcher([]))

    assert result.status == "ERROR"
    assert result.object_code == "ERROR"
    assert result.reason == ""
    assert "Thiếu mã đối tượng để sinh Lí do RPA" in result.error_note

    output_file = tmp_path / "rpa_input.xlsx"
    write_excel([result], output_file)
    wb = load_workbook(output_file)
    assert wb["BAO_NO_INPUT"].max_row == 1
    assert wb["BAO_CO_INPUT"].max_row == 1
    assert wb["EXCEPTION"].max_row == 2


def test_output_has_separate_flow_sheets_and_excludes_error(tmp_path):
    bao_no = ProcessedTransaction(
        source_file="sample.xlsx",
        original_row_index=2,
        bank="ACB",
        flow="bao_no",
        transaction_date=date(2026, 4, 1),
        object_code="ABC",
        object_name="ABC",
        reason="Thanh toán ABC",
        debit_account="331",
        credit_account="1121CT",
        amount=1000,
        use_case="Chi phí thanh toán",
        original_content="TT CHO ABC",
        counterparty_raw="",
        doc_no="1",
        status="OK",
        error_note="",
        confidence=0.95,
    )
    bao_co = ProcessedTransaction(
        source_file="sample.xlsx",
        original_row_index=3,
        bank="VCB",
        flow="bao_co",
        transaction_date=date(2026, 4, 1),
        object_code="",
        object_name="",
        reason="Lãi tiền gửi",
        debit_account="1121VCB",
        credit_account="515",
        amount=1000,
        use_case="Doanh thu tài chính",
        original_content="LAI NHAP VON",
        counterparty_raw="",
        doc_no="2",
        status="OK",
        error_note="",
        confidence=0.95,
    )
    err = ProcessedTransaction(
        source_file="sample.xlsx",
        original_row_index=4,
        bank="VCB",
        flow="bao_no",
        transaction_date=date(2026, 4, 2),
        object_code="ERROR",
        object_name="",
        reason="",
        debit_account="",
        credit_account="1121VCB",
        amount=2000,
        use_case="",
        original_content="BHXH",
        counterparty_raw="",
        doc_no="3",
        status="ERROR",
        error_note="Luồng bảo hiểm không xử lý tự động",
        confidence=0.99,
    )
    output_file = tmp_path / "rpa_input.xlsx"
    write_excel([bao_no, bao_co, err], output_file)
    wb = load_workbook(output_file)
    assert "BAO_NO_INPUT" in wb.sheetnames
    assert "BAO_CO_INPUT" in wb.sheetnames
    assert "THU_TIEN_MAT_INPUT" in wb.sheetnames
    assert "CHI_TIEN_MAT_INPUT" in wb.sheetnames
    assert "RPA_INPUT" not in wb.sheetnames
    assert [cell.value for cell in wb["BAO_NO_INPUT"][1]] == RPA_COLUMNS
    assert [cell.value for cell in wb["BAO_CO_INPUT"][1]] == RPA_BUSINESS_COLUMNS
    assert len([cell.value for cell in wb["BAO_NO_INPUT"][1]]) == len(RPA_COLUMNS)
    bao_no_values = dict(zip([cell.value for cell in wb["BAO_NO_INPUT"][1]], [cell.value for cell in wb["BAO_NO_INPUT"][2]]))
    assert bao_no_values["Ngân hàng"] == "ACB"
    assert wb["BAO_NO_INPUT"].max_row == 2
    assert wb["BAO_CO_INPUT"].max_row == 2


def test_write_object_match_review_for_object_errors(tmp_path):
    err = ProcessedTransaction(
        source_file="sample.xlsx",
        original_row_index=4,
        bank="VCB",
        flow="bao_no",
        transaction_date=date(2026, 4, 2),
        object_code="ERROR",
        object_name="",
        reason="",
        debit_account="331",
        credit_account="1121VCB",
        amount=2000,
        use_case="Chi phí thanh toán",
        original_content="TT CHO XI MANG SONG LAM",
        counterparty_raw="",
        doc_no="3",
        status="ERROR",
        error_note="Không tìm thấy mã đối tượng",
        confidence=0.79,
        matched_candidates=[
            ObjectCandidate(code="XMSONGLAM", name="Công ty cổ phần xi măng Sông Lam", score=79.5, source="fuzzy_name", matched_on="XI MANG SONG LAM")
        ],
        entities=ExtractedEntities(counterparty_hint="XI MANG SONG LAM"),
    )
    ambiguous = ProcessedTransaction(
        source_file="sample.xlsx",
        original_row_index=5,
        bank="ACB",
        flow="bao_no",
        transaction_date=date(2026, 4, 2),
        object_code="ERROR",
        object_name="",
        reason="",
        debit_account="331",
        credit_account="1121CT",
        amount=3000,
        use_case="Chi phí thanh toán",
        original_content="TT CHO QUANG MINH",
        counterparty_raw="",
        doc_no="4",
        status="ERROR",
        error_note="Nhiều mã đối tượng khớp gần bằng nhau",
        confidence=0.9,
        matched_candidates=[
            ObjectCandidate(code="QUANGMINH", name="Công ty Quang Minh", score=100, source="entity_match", matched_on="QUANG MINH"),
            ObjectCandidate(code="DLOC", name="Công ty Quang Minh Đại Lộc", score=100, source="entity_match", matched_on="QUANG MINH DAI LOC"),
        ],
        entities=ExtractedEntities(counterparty_hint="QUANG MINH"),
    )
    output_file = tmp_path / "object_match_review.xlsx"
    write_object_match_review(
        [err, ambiguous],
        output_file,
        alias_audit_rows=[
            {
                "catalog": "payable",
                "code": "VINACONTROL HP",
                "alias": "VINACONTROL",
                "risk": "unsafe_collision",
                "collision_count": 2,
                "hit_codes": "VINACONTROL HP, VINACONTROL QN",
                "hit_names": "Vinacontrol Hai Phong | Vinacontrol Quang Ninh",
            }
        ],
    )
    wb = load_workbook(output_file)
    assert "OBJECT_ERRORS" in wb.sheetnames
    assert "HINT_COLLISIONS" in wb.sheetnames
    assert "ALIAS_RISK" in wb.sheetnames
    assert "OBJECT_ACTIONS" in wb.sheetnames
    headers = [cell.value for cell in wb["OBJECT_ERRORS"][1]]
    values = dict(zip(headers, [cell.value for cell in wb["OBJECT_ERRORS"][2]]))
    assert values["reason_class"] == "missing_alias_or_low_score"
    assert values["best_candidate_code"] == "XMSONGLAM"
    action_headers = [cell.value for cell in wb["OBJECT_ACTIONS"][1]]
    action_types = [
        dict(zip(action_headers, [cell.value for cell in row]))["action_type"]
        for row in wb["OBJECT_ACTIONS"].iter_rows(min_row=2)
    ]
    assert "remove_or_narrow_alias" in action_types
    assert "review_ambiguous_hint" in action_types
    assert "review_alias_candidate" in action_types
