from __future__ import annotations

from datetime import date
from functools import lru_cache
from pathlib import Path

from openpyxl import load_workbook

from src.config_loader import load_config, load_rules
from src.entity_extractor import EntityExtractor, OwnCompanyConfig
from src.flows import FLOW_BAO_CO, FLOW_BAO_NO, FLOW_CHI_TIEN_MAT, FLOW_THU_TIEN_MAT
from src.models import Transaction
from src.object_aliases import load_object_aliases
from src.object_matcher import ObjectMatcher
from src.object_overrides import load_object_overrides
from src.output_writer import RPA_BUSINESS_COLUMNS, RPA_THU_TIEN_MAT_COLUMNS, write_excel, write_outputs
from src.parsers.acb_parser import ACBParser
from src.parsers.msb_parser import MSBParser
from src.parsers.vcb_parser import VCBParser
from src.processor import process_all, process_transaction
from src.rule_engine import RuleEngine


PROJECT_ROOT = Path(__file__).resolve().parents[1]
STATEMENTS_DIR = PROJECT_ROOT / "input" / "statements"
INTERNAL_RECORDS = [
    {"code": "DUC", "name": "Lê Ngọc Đức"},
    {"code": "HOA", "name": "Lê Thị Thanh Hoa"},
    {"code": "VIETHUNG", "name": "Phạm Việt Hùng"},
]
INTERNAL_ALIASES = {
    "DUC": ["LE NGOC DUC", "NGOC DUC"],
    "HOA": ["LE THI THANH HOA", "THANH HOA"],
    "VIETHUNG": ["PHAM VIET HUNG", "VIET HUNG"],
}


def _statement_sample(filename: str) -> Path:
    for base in [STATEMENTS_DIR, PROJECT_ROOT / "input"]:
        candidate = base / filename
        if candidate.exists():
            return candidate
    return STATEMENTS_DIR / filename


def _config() -> dict:
    config = load_config(PROJECT_ROOT / "config" / "config.yaml")
    config["ml"]["enabled"] = False
    return config


def _engine() -> RuleEngine:
    rules, _ = load_rules(None, PROJECT_ROOT / "config" / "default_rules.yaml")
    return RuleEngine(rules)


def _txn(description: str, debit: float = 0, credit: float = 0, bank: str = "VCB", counterparty: str = "") -> Transaction:
    return Transaction(
        source_file="sample.xlsx",
        bank=bank,
        transaction_date=date(2026, 4, 1),
        doc_no="REF1",
        description=description,
        counterparty_raw=counterparty,
        debit_amount=debit,
        credit_amount=credit,
        original_row_index=2,
    )


def _process(
    description: str,
    debit: float = 0,
    credit: float = 0,
    bank: str = "VCB",
    counterparty: str = "",
    receivable: list[dict] | None = None,
    payable: list[dict] | None = None,
    internal: list[dict] | None = None,
):
    return process_transaction(
        _txn(description, debit=debit, credit=credit, bank=bank, counterparty=counterparty),
        _config(),
        _engine(),
        ObjectMatcher.from_records(receivable or []),
        ObjectMatcher.from_records(payable or []),
        internal_matcher=ObjectMatcher.from_records(internal or INTERNAL_RECORDS, aliases=INTERNAL_ALIASES),
    )


@lru_cache(maxsize=1)
def _real_processing_context():
    own_company = OwnCompanyConfig.from_yaml(PROJECT_ROOT / "config" / "own_company.yaml")
    aliases = load_object_aliases(PROJECT_ROOT / "config" / "object_aliases.yaml")
    overrides = load_object_overrides(PROJECT_ROOT / "config" / "object_overrides.yaml")

    def merge_aliases(catalog: str) -> dict[str, list[str]]:
        merged: dict[str, list[str]] = {}
        for section in [aliases.get(catalog, {}), overrides.get(catalog, {}).get("aliases", {})]:
            for code, values in section.items():
                merged.setdefault(code, []).extend(values or [])
        return {code: list(dict.fromkeys(values)) for code, values in merged.items()}

    receivable = ObjectMatcher.from_excel(
        PROJECT_ROOT / "input" / "R_DMDT1 1.xlsx",
        aliases=merge_aliases("receivable"),
        exact_phrase_overrides=overrides.get("receivable", {}).get("exact_phrases", {}),
        supplemental_objects=overrides.get("receivable", {}).get("supplemental_objects", []),
        own_company=own_company,
    )
    payable = ObjectMatcher.from_excel(
        PROJECT_ROOT / "input" / "R_DMDT1.xlsx",
        aliases=merge_aliases("payable"),
        exact_phrase_overrides=overrides.get("payable", {}).get("exact_phrases", {}),
        supplemental_objects=overrides.get("payable", {}).get("supplemental_objects", []),
        own_company=own_company,
    )
    internal = ObjectMatcher.from_excel(PROJECT_ROOT / "input" / "MA NOI BO CTY.xlsx", own_company=own_company)
    return receivable, payable, internal, EntityExtractor(own_company)


def _process_real(description: str, debit: float = 0, credit: float = 0, bank: str = "ACB"):
    receivable, payable, internal, extractor = _real_processing_context()
    return process_transaction(
        _txn(description, debit=debit, credit=credit, bank=bank),
        _config(),
        _engine(),
        receivable,
        payable,
        internal_matcher=internal,
        entity_extractor=extractor,
    )


def test_parsers_read_real_acb_vcb_msb_and_skip_msb_totals():
    acb = ACBParser().parse(_statement_sample("5614249_SAOKE_TK_202604 (2).xlsx"))
    vcb = VCBParser().parse(_statement_sample("lich-su-giao-dich-tai-khoan VCB T4.26.xls"))
    msb_parser = MSBParser()
    msb = msb_parser.parse(_statement_sample("ReportIBSCorpAccountStatement_20260526165427.xlsx"))

    assert any(item.debit_amount > 0 for item in acb)
    assert any(item.credit_amount > 0 for item in acb)
    assert any(item.debit_amount > 0 for item in vcb)
    assert any(item.credit_amount > 0 for item in vcb)
    assert any(item.debit_amount > 0 for item in msb)
    assert any(item.credit_amount > 0 for item in msb)
    assert not any(item.original_row_index >= 118 for item in msb)
    assert msb_parser.skipped_row_count >= 3


def test_bao_co_rules_use_specific_accounts_before_customer_receivable():
    customer = [{"code": "ABC", "name": "Cong ty ABC"}]
    assert _process("ABC THANH TOAN CONG NO HD 123", credit=100, counterparty="ABC", receivable=customer).credit_account == "131"
    assert _process("LAI NHAP VON", credit=100).credit_account == "515"
    assert _process("CREDIT INTEREST", credit=100).credit_account == "515"
    assert _process("TRA LAI TAI KHOAN", credit=100).credit_account == "515"
    advance_refund = _process("HOAN LAI TAM UNG", credit=100, counterparty="LE NGOC DUC")
    assert advance_refund.credit_account == "141"
    assert advance_refund.object_code == "DUC"
    assert _process("BAN NGOAI TE TY GIA USD VND", credit=100).credit_account == "1122"
    assert _process("MUA TU BAO CO NGOAI TE THANH TOAN HD", credit=100, receivable=customer).credit_account == "1122"
    assert _process("GIAI NGAN KHOAN VAY", credit=100).credit_account == "341"
    assert _process("LE NGOC DUC HOAN VAY", credit=100).credit_account != "341"
    fx1 = _process("M1HH/KHDN/ MUA TU BAO CO SO TIEN 42000 USD, TY GIA 26.217", credit=1101114000, bank="ACB")
    assert fx1.status == "OK"
    assert fx1.credit_account == "1122"
    assert fx1.foreign_currency == "USD"
    assert fx1.foreign_amount == 42000
    assert fx1.exchange_rate == 26217
    assert fx1.reason == "Bán ngoại tệ 42000 USD tỷ giá 26217"
    fx2 = _process("M1HH/KHDN/ MUA TU BAO CO KH SO TIEN 50.000 USD, TY GIA 26228,", credit=1311400000, bank="ACB")
    assert fx2.status == "OK"
    assert fx2.credit_account == "1122"
    assert fx2.foreign_amount == 50000
    assert fx2.exchange_rate == 26228


def test_bao_no_rules_for_forex_loan_and_existing_bank_fee():
    assert _process("MUA NGOAI TE", debit=100).debit_account == "1122"
    assert _process("MUA USD", debit=100).debit_account == "1122"
    assert _process("THU NO TK VAY 001065887769", debit=100).debit_account == "341"
    assert _process("TRA GOC KHOAN VAY", debit=100).debit_account == "341"
    assert _process("THANH TOAN USD", debit=100).debit_account != "1122"
    assert _process("PHI NGAN HANG", debit=100).debit_account == "635"
    assert _process("THU PHI PHAT HANH BAO LANH THUC HIEN HOP DONG", debit=100, bank="ACB").debit_account == "635"
    assert _process("NOP THUE GTGT THANG 4", debit=100).debit_account == "3331"


def test_acb_exception_patterns_pass_with_approved_overrides():
    receivable_cases = [
        ("CTY NGUYEN KIM 0109912477 STSTMSHP 2604109 GD 6125IBT1FJQI1KNQ", "NGUYENKIM"),
        ("PIL VIETNAM CO LTD TAX CODE - 0303449450-[3186493995] NHTMCP A CHAU HCM HCM PIL PAY INV 584", "PIL VN"),
        ("CTY TNHH PTXD VA TM 0101101999 PHI PHAT LENH SO BL STSTMSHP2604102", "PTXDVATM"),
        ("TCL SMART DEVICE VIET NAM COMPANY LIMITED-TIEN COC", "THONGMINHTCL"),
        ("CONG TY MINH HUY THANH TON TIEN THUE VAN PHONG VA DIEN NUOC THANG 02 03", "MINHHUY"),
        ("MBVCB.14166421961.394500.CTY SAKAI NOP PHI BILL 2603107", "SAKAI"),
        ("CP VIETNAM CORPORATION-553 VOI HOT-TA 25KGBAO CHARGEDETAILS OUR", "CPVIETNAM"),
        ("CN HN-CT TNHH MINH PHONG HOP NHAT CHUYEN TIEN DIEN NUOC T3/2026", "MINHPHONGHOPNHAT"),
        ("CO SO THANH PHONG CHUYEN TIEN GD 6135IBT1CJ1HNHAG", "THANHPHONG"),
        ("CHI NHANH CONG TY TNHH DO SUNG MACHINERY CHUYEN TIEN MUA KET SAT CHO CTY LE PHAM", "DOSUNG"),
    ]
    for description, object_code in receivable_cases:
        result = _process_real(description, credit=100)
        assert result.status == "OK"
        assert result.flow == FLOW_BAO_CO
        assert result.credit_account == "131"
        assert result.object_code == object_code

    payable_cases = [
        ("CT TNHH LE PHAM TT TIEN HANG CHO CT VIET THANG", "VIETTHAG"),
        ("LE PHAM CT CHO THIEN SON", "THIENSON"),
        ("LE PHAM CT CHO VOI VIET", "VOIVIET"),
        ("LE PHAM CT CHO 268", "26868"),
        ("TT TIENHANG CHO CT CP HC MINH DUC", "MD"),
        ("CT TNHH LE PHAM TT PHI GIAM DINH CHO SGS VIET NAM", "SGS"),
        ("LE PHAM DAT COC THIET KE PHAN MEM TU DONG HOA NGHIEP VU", "PHANMEM"),
        ("CTY LEPHAM TT CUOC BIEN CHO VSICO", "VSICO"),
        ("LE PHAM CT CHO NHAT MINH", "VTBNHATMINH"),
        ("CT TNHH LE PHAM TT CANG GAMA. TAU EAGLE AROW, ST 22376.67, TG 26125.", "CANGGAMA"),
    ]
    for description, object_code in payable_cases:
        result = _process_real(description, debit=100)
        assert result.status == "OK"
        assert result.flow == FLOW_BAO_NO
        assert result.debit_account == "331"
        assert result.object_code == object_code


def test_bill_issue_fee_receipts_use_customer_001_and_company_reason():
    cases = [
        (
            "CTY TNHH PTXD VA TM 0101101999 PHI PHAT LENH SO BL STSTMSHP2604102",
            "CTY TNHH PTXD VA TM",
        ),
        (
            "MBVCB.14166421961.394500.CTY SAKAI NOP PHI BILL 2603107",
            "CTY SAKAI",
        ),
    ]
    for description, company_name in cases:
        result = _process_real(description, credit=1100000, bank="MSB")
        assert result.status == "OK"
        assert result.flow == FLOW_BAO_CO
        assert result.credit_account == "131"
        assert result.object_code == "001"
        assert result.object_name == company_name
        assert result.reason == f"Thanh toán phí phát lệnh ({company_name})"


def test_bill_issue_fee_receipt_requires_amount_guard():
    result = _process_real(
        "CTY CLIO SHIPPING TT PHI DO BL 106043ZJ001 DEN ZJ009 MV LILA NOLA",
        credit=4860000,
    )
    assert result.status == "OK"
    assert result.object_code == "CLIO SHIPPING"

    same_amount_without_fee_signal = _process_real(
        "CTY NGUYEN KIM 0109912477 STSTMSHP 2604109 GD 6125IBT1FJQI1KNQ",
        credit=1100000,
    )
    assert same_amount_without_fee_signal.status == "OK"
    assert same_amount_without_fee_signal.object_code == "NGUYENKIM"


def test_company_advance_fee_uses_payable_331():
    advance = _process_real("T UNG CHO HOANG ANH TT PHI NANG HA CHO CT CP CANG HAI PHONG", debit=100)
    assert advance.status == "OK"
    assert advance.debit_account == "331"
    assert advance.object_code == "CANG-HP"


def test_acb_negative_patterns_stay_manual_review():
    assert _process_real("CHUYEN TIEN GD 6140IBT1FJW5PHSH", credit=100).status == "ERROR"
    own_transfer = _process_real("CONG TY TNHH LE PHAM CHUYEN TK GD 6141IBT1FJW4DKQJ", credit=100)
    assert own_transfer.status == "ERROR"
    assert own_transfer.credit_account != "131"


def test_cash_receipt_and_cash_payment_are_exclusive_flows():
    cash_in = _process("RUT TIEN MAT NHAP QUY", debit=100, bank="ACB")
    assert cash_in.flow == FLOW_THU_TIEN_MAT
    assert cash_in.debit_account == "1111"
    assert cash_in.credit_account == "1121CT"

    cheque_cash_in = _process("LE THI THANH HOA#001178043963#CHI SEC 22541668#1156992 ;", debit=3000000000, bank="ACB")
    assert cheque_cash_in.flow == FLOW_THU_TIEN_MAT
    assert cheque_cash_in.debit_account == "1111"
    assert cheque_cash_in.credit_account == "1121CT"
    assert cheque_cash_in.entities.cash_person_name == "LE THI THANH HOA"

    fee = _process("PHI RUT TIEN", debit=100, bank="VCB")
    assert fee.flow != FLOW_THU_TIEN_MAT

    cash_out = _process("NOP TIEN MAT VAO TAI KHOAN", credit=100, bank="MSB")
    assert cash_out.flow == FLOW_CHI_TIEN_MAT
    assert cash_out.debit_account == "1121HB"
    assert cash_out.credit_account == "1111"

    acb_cash_out = _process("LE THI THANH HOA#001178043963#NT#12739876;12739876-NT-TK-ACB-1794027", credit=100, bank="ACB")
    assert acb_cash_out.flow == FLOW_CHI_TIEN_MAT
    assert acb_cash_out.debit_account == "1121CT"
    assert acb_cash_out.credit_account == "1111"
    assert acb_cash_out.entities.cash_person_name == "LE THI THANH HOA"

    customer = [{"code": "KHACHHANG", "name": "Khach Hang"}]
    receivable = _process(
        "KHACH HANG NOP TIEN THANH TOAN HOA DON",
        credit=100,
        counterparty="KHACH HANG",
        receivable=customer,
    )
    assert receivable.flow == FLOW_BAO_CO
    assert receivable.credit_account == "131"


def test_personal_advance_to_company_goes_bao_co_141_not_receivable():
    cases = [
        "LE NGOC DUC CHUYEN TIEN GD 6124IBT1AWZTW1E7 040526-14:04:19",
        "LE NGOC DUC CHUYEN TIEN GD 6124IBT1AWZTW2S9 040526-14:04:55",
        "IB LE NGOC DUC CHUYEN KHOAN",
    ]
    for description in cases:
        result = _process(description, credit=100, bank="ACB")
        assert result.status == "OK"
        assert result.flow == FLOW_BAO_CO
        assert result.debit_account == "1121CT"
        assert result.credit_account == "141"
        assert result.object_code == "DUC"
        assert result.reason == "Nhận tiền tạm ứng cá nhân DUC"
        assert result.matched_rule == "personal_advance_to_company"


def test_advance_splits_company_payable_and_internal_person():
    company = _process(
        "TAM UNG CHO CONG TY ABC",
        debit=100,
        bank="ACB",
        payable=[{"code": "ABC", "name": "Cong ty ABC"}],
    )
    assert company.status == "OK"
    assert company.flow == FLOW_BAO_NO
    assert company.debit_account == "331"
    assert company.object_code == "ABC"
    assert company.reason == "Tạm ứng tiền hàng ABC"

    person = _process("TAM UNG CHO LE NGOC DUC", debit=100, bank="ACB")
    assert person.status == "OK"
    assert person.flow == FLOW_BAO_NO
    assert person.debit_account == "141"
    assert person.object_code == "DUC"
    assert person.reason == "Tạm ứng cá nhân DUC"

    alias_person = _process("TAM UNG CHO VIET HUNG", debit=100, bank="ACB")
    assert alias_person.status == "OK"
    assert alias_person.flow == FLOW_BAO_NO
    assert alias_person.debit_account == "141"
    assert alias_person.object_code == "VIETHUNG"


def test_rule_first_does_not_call_ml_when_rule_matched():
    class RaisingClassifier:
        def predict(self, *args, **kwargs):  # noqa: ANN002, ANN003
            raise AssertionError("ML should not be called after a rule match")

    result = process_transaction(
        _txn("TRA LAI TAI KHOAN", credit=100),
        _config(),
        _engine(),
        ObjectMatcher([]),
        ObjectMatcher([]),
        classifier=RaisingClassifier(),
    )
    assert result.status == "OK"
    assert result.flow == FLOW_BAO_CO
    assert result.credit_account == "515"


def test_output_has_four_pad_sheets_with_exchange_rate_column(tmp_path):
    items = [
        _process("RUT TIEN MAT NHAP QUY", debit=100, bank="ACB"),
        _process("NOP TIEN MAT VAO TAI KHOAN", credit=200, bank="VCB"),
        _process("MUA USD", debit=300, bank="MSB"),
        _process("TRA LAI TAI KHOAN", credit=400, bank="MSB"),
    ]
    output_file = tmp_path / "rpa_input.xlsx"
    write_excel(items, output_file, run_id="run1", run_stats={"skipped_non_transaction_rows": 3})
    wb = load_workbook(output_file, data_only=True)

    for sheet_name in ["BAO_NO_INPUT", "BAO_CO_INPUT", "THU_TIEN_MAT_INPUT", "CHI_TIEN_MAT_INPUT"]:
        assert sheet_name in wb.sheetnames
        expected_columns = RPA_THU_TIEN_MAT_COLUMNS if sheet_name == "THU_TIEN_MAT_INPUT" else RPA_BUSINESS_COLUMNS
        assert [cell.value for cell in wb[sheet_name][1]] == expected_columns
        assert wb[sheet_name].max_column == len(expected_columns)
    assert "RPA_TASKS" in wb.sheetnames
    assert "AUDIT_LOG" in wb.sheetnames
    assert "MANUAL_REVIEW" in wb.sheetnames


def test_rpa_input_exports_foreign_exchange_rate(tmp_path):
    bao_co = _process(
        "M1HH/KHDN/ MUA TU BAO CO SO TIEN 42000 USD, TY GIA 26.217",
        credit=1101114000,
        bank="ACB",
    )
    bao_no = _process(
        "MUA NGOAI TE SO TIEN 1000 USD, TY GIA 25.100",
        debit=25100000,
        bank="ACB",
    )
    output_file = tmp_path / "rpa_input.xlsx"
    write_excel([bao_co, bao_no], output_file)
    wb = load_workbook(output_file, data_only=True)

    bao_co_headers = [cell.value for cell in wb["BAO_CO_INPUT"][1]]
    bao_co_values = dict(zip(bao_co_headers, [cell.value for cell in wb["BAO_CO_INPUT"][2]]))
    assert bao_co_values["Tỷ giá"] == 26217

    bao_no_headers = [cell.value for cell in wb["BAO_NO_INPUT"][1]]
    bao_no_values = dict(zip(bao_no_headers, [cell.value for cell in wb["BAO_NO_INPUT"][2]]))
    assert bao_no_values["Tỷ giá"] == 25100


def test_cash_flows_export_recipient_name(tmp_path):
    cash_in = _process("LE THI THANH HOA#001178043963#CHI SEC 22541668#1156992 ;", debit=3000000000, bank="ACB")
    cash_out = _process("LE THI THANH HOA#001178043963#NT#12739876;12739876-NT-TK-ACB-1794027", credit=100, bank="ACB")
    output_file = tmp_path / "rpa_input.xlsx"
    write_excel([cash_in, cash_out], output_file)
    wb = load_workbook(output_file, data_only=True)

    thu_headers = [cell.value for cell in wb["THU_TIEN_MAT_INPUT"][1]]
    thu_values = dict(zip(thu_headers, [cell.value for cell in wb["THU_TIEN_MAT_INPUT"][2]]))
    assert thu_values["Người nộp tiền"] == "LE THI THANH HOA"

    chi_headers = [cell.value for cell in wb["CHI_TIEN_MAT_INPUT"][1]]
    chi_values = dict(zip(chi_headers, [cell.value for cell in wb["CHI_TIEN_MAT_INPUT"][2]]))
    assert chi_values["Người nhận tiền"] == "LE THI THANH HOA"


def test_integration_process_real_samples_and_write_outputs(tmp_path):
    import logging

    config = _config()
    logger = logging.getLogger("integration-four-flows")
    processed = process_all(
        statements_dir=STATEMENTS_DIR,
        receivable_path=PROJECT_ROOT / "input" / "R_DMDT1 1.xlsx",
        payable_path=PROJECT_ROOT / "input" / "R_DMDT1.xlsx",
        rules_path=None,
        default_rules_path=PROJECT_ROOT / "config" / "default_rules.yaml",
        config=config,
        logger=logger,
    )
    result = write_outputs(processed, tmp_path, config)
    wb = load_workbook(result.excel_path, data_only=True)

    assert len(processed) > 0
    assert not any(item.bank == "MSB" and item.original_row_index >= 118 for item in processed)
    for sheet_name in ["BAO_NO_INPUT", "BAO_CO_INPUT", "THU_TIEN_MAT_INPUT", "CHI_TIEN_MAT_INPUT"]:
        assert sheet_name in wb.sheetnames
        expected_columns = RPA_THU_TIEN_MAT_COLUMNS if sheet_name == "THU_TIEN_MAT_INPUT" else RPA_BUSINESS_COLUMNS
        assert [cell.value for cell in wb[sheet_name][1]] == expected_columns
    assert "SUMMARY" in wb.sheetnames
    assert "RPA_TASKS" in wb.sheetnames
