from __future__ import annotations

from dataclasses import asdict, dataclass
from datetime import date
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from .flows import (
    FLOW_CHI_TIEN_MAT,
    FLOW_SPECS,
    FLOW_THU_TIEN_MAT,
    PAD_FLOW_ORDER,
    flow_execution_order,
    flow_sheet,
    flow_voucher_type,
)
from .models import ProcessedTransaction
from .rpa_input_status import (
    INPUT_MESSAGE_COLUMN,
    INPUT_STATUS_COLUMN,
    INPUT_UPDATED_AT_COLUMN,
)
from .rpa_summary import RpaRunState, prepare_rpa_run, write_summary
from .rpa_tracking import STATUS_PENDING, write_tracking_records
from .vietnamese_encoding import unicode_to_tcvn3


RPA_INPUT_STATUS_COLUMNS = [
    INPUT_STATUS_COLUMN,
    INPUT_MESSAGE_COLUMN,
    INPUT_UPDATED_AT_COLUMN,
]
RPA_BUSINESS_COLUMNS = [
    "Ngày CT",
    "Mã ĐT",
    "Lí do",
    "Người nhận tiền",
    "TK nợ",
    "TK có",
    "Thành tiền",
    "Tỷ giá",
    "Ngân hàng",
    "transaction_uid",
    "run_id",
] + RPA_INPUT_STATUS_COLUMNS
RPA_THU_TIEN_MAT_COLUMNS = [
    "Ngày CT",
    "Mã ĐT",
    "Lí do",
    "Người nộp tiền",
    "TK nợ",
    "TK có",
    "Thành tiền",
    "Tỷ giá",
    "Ngân hàng",
    "transaction_uid",
    "run_id",
] + RPA_INPUT_STATUS_COLUMNS
RPA_TRACKING_COLUMNS = [
    "transaction_uid",
    "rpa_status",
    "rpa_message",
    "source_file",
    "source_sheet",
    "source_row",
    "direction",
]
RPA_COLUMNS = RPA_BUSINESS_COLUMNS
RPA_REASON_ENCODING_TCVN3 = "tcvn3"
RPA_REASON_UNICODE_COLUMN = "Lí do Unicode"
RPA_TASK_COLUMNS = [
    "run_id",
    "task_id",
    "transaction_uid",
    "flow",
    "input_sheet",
    "input_excel_row",
    "summary_status",
    "source_file",
    "source_sheet",
    "source_row_index",
    "sheet_row_count",
    "execution_order",
    "voucher_type",
    "has_error",
]


@dataclass
class OutputWriteResult:
    run_id: str
    excel_path: Path
    tracking_path: Path | None
    summary_path: Path
    object_match_review_path: Path
    stats: dict[str, int]


def write_outputs(
    processed: list[ProcessedTransaction],
    output_dir: str | Path,
    config: dict[str, Any],
    logger: Any | None = None,
) -> OutputWriteResult:
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    output_cfg = config.get("output", {})
    excel_path = output_dir / output_cfg.get("excel_file", "rpa_input.xlsx")
    summary_path = output_dir / output_cfg.get("summary_file", "rpa_summary.xlsx")
    object_match_review_path = output_dir / output_cfg.get("object_match_review_file", "object_match_review.xlsx")
    run_state = prepare_rpa_run(processed, summary_path, logger=logger)
    write_summary(run_state.summary_df, summary_path)
    write_excel(
        processed,
        excel_path,
        run_id=run_state.run_id,
        rpa_items=run_state.rpa_items,
        run_stats=config.get("_run_stats", {}),
        rpa_reason_encoding=output_cfg.get("rpa_reason_encoding", ""),
    )
    write_object_match_review(processed, object_match_review_path, alias_audit_rows=_build_alias_audit_rows(config))
    return OutputWriteResult(
        run_id=run_state.run_id,
        excel_path=excel_path,
        tracking_path=None,
        summary_path=summary_path,
        object_match_review_path=object_match_review_path,
        stats=run_state.stats,
    )


def write_excel(
    processed: list[ProcessedTransaction],
    path: str | Path,
    run_id: str | None = None,
    rpa_items: list[ProcessedTransaction] | None = None,
    run_stats: dict[str, Any] | None = None,
    rpa_reason_encoding: str = "",
) -> None:
    path = Path(path)
    input_items = rpa_items if rpa_items is not None else processed
    flow_items = {
        flow: [item for item in input_items if item.status == "OK" and item.flow == flow]
        for flow in PAD_FLOW_ORDER
    }
    flow_dfs = {
        flow: pd.DataFrame(
            [_rpa_record(item, run_id=run_id, reason_encoding=rpa_reason_encoding) for item in items],
            columns=_rpa_columns_for_flow(flow, rpa_reason_encoding),
        )
        for flow, items in flow_items.items()
    }
    exception_df = pd.DataFrame([_exception_record(item) for item in processed if item.status != "OK"])
    manual_review_df = pd.DataFrame([_manual_review_record(item) for item in processed if item.status != "OK"])
    audit_df = pd.DataFrame([_audit_record(item) for item in processed])
    parser_warnings_df = pd.DataFrame((run_stats or {}).get("parser_warnings", []))
    summary_df = pd.DataFrame(_summary_records(processed, flow_items, run_stats or {}))
    task_df = pd.DataFrame(_task_records(flow_items, run_id), columns=RPA_TASK_COLUMNS)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for flow in PAD_FLOW_ORDER:
            flow_dfs[flow].to_excel(writer, sheet_name=flow_sheet(flow), index=False)
        exception_df.to_excel(writer, sheet_name="EXCEPTION", index=False)
        manual_review_df.to_excel(writer, sheet_name="MANUAL_REVIEW", index=False)
        audit_df.to_excel(writer, sheet_name="AUDIT_LOG", index=False)
        parser_warnings_df.to_excel(writer, sheet_name="PARSER_WARNINGS", index=False)
        summary_df.to_excel(writer, sheet_name="SUMMARY", index=False)
        if run_id:
            task_df.to_excel(writer, sheet_name="RPA_TASKS", index=False)
        sheet_names = [flow_sheet(flow) for flow in PAD_FLOW_ORDER] + [
            "EXCEPTION",
            "MANUAL_REVIEW",
            "AUDIT_LOG",
            "PARSER_WARNINGS",
            "SUMMARY",
        ]
        if run_id:
            sheet_names.append("RPA_TASKS")
        for sheet_name in sheet_names:
            _format_sheet(writer.book[sheet_name])


def write_tracking(processed: list[ProcessedTransaction], path: str | Path, run_state: RpaRunState) -> None:
    processed_by_uid = {item.transaction_uid: item for item in processed if item.transaction_uid}
    records: list[dict[str, Any]] = []
    for summary_record in run_state.summary_df.to_dict("records"):
        uid = str(summary_record.get("transaction_uid", "")).strip()
        if not uid:
            continue
        item = processed_by_uid.get(uid)
        if item:
            record = _tracking_record(item)
            _apply_summary_tracking_fields(record, summary_record)
        else:
            record = _tracking_record_from_summary(summary_record)
        records.append(record)
    write_tracking_records(records, path)


OBJECT_MATCH_REVIEW_COLUMNS = [
    "reason_class",
    "suggested_action",
    "error_note",
    "source_file",
    "source_sheet",
    "source_row",
    "bank",
    "flow",
    "use_case",
    "original_content",
    "counterparty_raw",
    "counterparty_hint",
    "counterparty_source",
    "tax_code",
    "object_match_source",
    "best_candidate_code",
    "best_candidate_name",
    "best_candidate_score",
    "best_candidate_source",
    "best_candidate_matched_on",
    "object_ml_status",
    "object_ml_best_code",
    "object_ml_confidence",
    "object_ml_gap",
    "object_ml_decision",
    "object_ml_note",
    "top_candidates",
    "transaction_uid",
]


HINT_COLLISION_COLUMNS = [
    "risk_class",
    "source_file",
    "source_sheet",
    "source_row",
    "bank",
    "flow",
    "counterparty_hint",
    "counterparty_source",
    "object_match_source",
    "matched_object_code",
    "best_candidate_code",
    "best_candidate_score",
    "second_candidate_code",
    "second_candidate_score",
    "score_gap",
    "candidate_codes",
    "object_ml_status",
    "object_ml_best_code",
    "object_ml_confidence",
    "object_ml_gap",
    "object_ml_decision",
    "original_content",
    "error_note",
    "suggested_action",
    "transaction_uid",
]

ALIAS_RISK_COLUMNS = [
    "catalog",
    "code",
    "alias",
    "risk",
    "collision_count",
    "hit_codes",
    "hit_names",
]

OBJECT_ACTION_COLUMNS = [
    "priority",
    "action_type",
    "reason_class",
    "catalog",
    "code",
    "alias_or_hint",
    "tax_code",
    "count",
    "source_refs",
    "candidate_codes",
    "object_ml_status",
    "object_ml_confidence",
    "object_ml_gap",
    "details",
    "suggested_action",
]


def write_object_match_review(
    processed: list[ProcessedTransaction],
    path: str | Path,
    alias_audit_rows: list[dict[str, Any]] | None = None,
) -> None:
    path = Path(path)
    rows = [_object_match_review_record(item) for item in processed if _has_object_match_error(item)]
    detail_df = pd.DataFrame(rows, columns=OBJECT_MATCH_REVIEW_COLUMNS)
    summary_df = _object_match_review_summary(detail_df)
    hint_collision_df = pd.DataFrame(_hint_collision_records(processed), columns=HINT_COLLISION_COLUMNS)
    alias_risk_df = pd.DataFrame(alias_audit_rows or [], columns=ALIAS_RISK_COLUMNS)
    object_action_df = pd.DataFrame(
        _object_action_records(detail_df, alias_risk_df),
        columns=OBJECT_ACTION_COLUMNS,
    )

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        detail_df.to_excel(writer, sheet_name="OBJECT_ERRORS", index=False)
        summary_df.to_excel(writer, sheet_name="SUMMARY", index=False)
        hint_collision_df.to_excel(writer, sheet_name="HINT_COLLISIONS", index=False)
        alias_risk_df.to_excel(writer, sheet_name="ALIAS_RISK", index=False)
        object_action_df.to_excel(writer, sheet_name="OBJECT_ACTIONS", index=False)
        _format_sheet(writer.book["OBJECT_ERRORS"])
        _format_sheet(writer.book["SUMMARY"])
        _format_sheet(writer.book["HINT_COLLISIONS"])
        _format_sheet(writer.book["ALIAS_RISK"])
        _format_sheet(writer.book["OBJECT_ACTIONS"])


def _build_alias_audit_rows(config: dict[str, Any]) -> list[dict[str, Any]]:
    runtime_paths = config.get("_runtime_paths", {})
    aliases_path = runtime_paths.get("object_aliases_path")
    if not aliases_path:
        return []

    from .entity_extractor import OwnCompanyConfig
    from .object_aliases import load_object_aliases
    from .object_matcher import ObjectMatcher

    own_company = OwnCompanyConfig.from_yaml(runtime_paths.get("own_company_path"))
    aliases = load_object_aliases(aliases_path)
    rows: list[dict[str, Any]] = []
    for catalog, path_key in (("payable", "payable_path"), ("receivable", "receivable_path")):
        catalog_path = runtime_paths.get(path_key)
        if not catalog_path:
            continue
        try:
            matcher = ObjectMatcher.from_excel(
                catalog_path,
                aliases=aliases.get(catalog, {}),
                own_company=own_company,
            )
        except Exception:  # noqa: BLE001 - audit must not block RPA output
            continue
        rows.extend(matcher.alias_audit_records(catalog))
    return rows


def _object_action_records(detail_df: pd.DataFrame, alias_risk_df: pd.DataFrame) -> list[dict[str, Any]]:
    rows = _alias_action_records(alias_risk_df) + _detail_action_records(detail_df)
    rows.sort(key=lambda row: (int(row.get("priority") or 99), str(row.get("action_type") or ""), -int(row.get("count") or 0)))
    return rows


def _alias_action_records(alias_risk_df: pd.DataFrame) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    if alias_risk_df.empty:
        return rows
    for record in alias_risk_df.to_dict("records"):
        risk = str(record.get("risk") or "")
        if risk not in {"unsafe_collision", "weak_alias", "blocked_alias", "missing_code"}:
            continue
        priority = 1 if risk == "unsafe_collision" else 4 if risk in {"weak_alias", "blocked_alias"} else 2
        action_type = {
            "unsafe_collision": "remove_or_narrow_alias",
            "weak_alias": "review_weak_alias",
            "blocked_alias": "review_blocked_alias",
            "missing_code": "fix_alias_code",
        }.get(risk, "review_alias")
        rows.append(
            {
                "priority": priority,
                "action_type": action_type,
                "reason_class": risk,
                "catalog": _clean_report_value(record.get("catalog")),
                "code": _clean_report_value(record.get("code")),
                "alias_or_hint": _clean_report_value(record.get("alias")),
                "tax_code": "",
                "count": int(record.get("collision_count") or 0),
                "source_refs": "",
                "candidate_codes": _clean_report_value(record.get("hit_codes")),
                "object_ml_status": "",
                "object_ml_confidence": "",
                "object_ml_gap": "",
                "details": _clean_report_value(record.get("hit_names")),
                "suggested_action": _object_action_suggestion(action_type),
            }
        )
    return rows


def _detail_action_records(detail_df: pd.DataFrame) -> list[dict[str, Any]]:
    if detail_df.empty:
        return []
    grouped: dict[tuple[str, str, str, str, str], dict[str, Any]] = {}
    for record in detail_df.to_dict("records"):
        reason_class = _clean_report_value(record.get("reason_class"))
        action_type = _action_type_for_reason(reason_class)
        priority = _priority_for_action(action_type)
        hint = _clean_report_value(record.get("counterparty_hint"))
        tax_code = _clean_report_value(record.get("tax_code"))
        best_code = _clean_report_value(record.get("best_candidate_code"))
        alias_or_hint = tax_code if action_type == "update_catalog_tax_code" else hint or _clean_report_value(record.get("best_candidate_matched_on"))
        code = best_code if action_type in {"review_alias_candidate", "review_ambiguous_hint"} else ""
        key = (action_type, reason_class, _catalog_for_flow(record.get("flow")), code, alias_or_hint or tax_code)
        current = grouped.setdefault(
            key,
            {
                "priority": priority,
                "action_type": action_type,
                "reason_class": reason_class,
                "catalog": _catalog_for_flow(record.get("flow")),
                "code": code,
                "alias_or_hint": alias_or_hint,
                "tax_code": tax_code,
                "count": 0,
                "source_refs": [],
                "candidate_codes": _candidate_codes(record.get("top_candidates")),
                "object_ml_status": _clean_report_value(record.get("object_ml_status")),
                "object_ml_confidence": _clean_report_value(record.get("object_ml_confidence")),
                "object_ml_gap": _clean_report_value(record.get("object_ml_gap")),
                "details": _clean_report_value(record.get("original_content")),
                "suggested_action": _object_action_suggestion(action_type),
            },
        )
        current["count"] += 1
        ref = _source_ref(record)
        if ref and ref not in current["source_refs"]:
            current["source_refs"].append(ref)
        if not current["candidate_codes"]:
            current["candidate_codes"] = _candidate_codes(record.get("top_candidates"))
    rows: list[dict[str, Any]] = []
    for row in grouped.values():
        row = dict(row)
        row["source_refs"] = "; ".join(row["source_refs"][:10])
        rows.append(row)
    return rows


def _action_type_for_reason(reason_class: str) -> str:
    return {
        "ambiguous": "review_ambiguous_hint",
        "tax_code_not_in_catalog": "update_catalog_tax_code",
        "missing_alias_or_low_score": "review_alias_candidate",
        "weak_or_generic_hint": "improve_extractor",
        "no_extractable_object": "manual_object_review",
        "not_in_catalog_or_missing_alias": "review_catalog_or_alias",
    }.get(reason_class, "review_object_error")


def _priority_for_action(action_type: str) -> int:
    return {
        "remove_or_narrow_alias": 1,
        "fix_alias_code": 2,
        "review_ambiguous_hint": 2,
        "update_catalog_tax_code": 3,
        "review_alias_candidate": 4,
        "review_weak_alias": 4,
        "review_blocked_alias": 4,
        "improve_extractor": 5,
        "manual_object_review": 5,
        "review_catalog_or_alias": 5,
    }.get(action_type, 9)


def _object_action_suggestion(action_type: str) -> str:
    actions = {
        "remove_or_narrow_alias": "Tạm vô hiệu alias rộng; thay bằng alias cụ thể có MST, số tài khoản hoặc tên đủ phân biệt.",
        "fix_alias_code": "Alias đang trỏ đến mã không có trong danh mục; sửa mã hoặc xóa alias.",
        "review_ambiguous_hint": "Chọn mã đúng; chỉ thêm alias mạnh, không dùng hint ngắn đang đụng nhiều mã.",
        "update_catalog_tax_code": "Bổ sung MST vào đúng mã ĐT trong danh mục hoặc tạo mã ĐT nếu đối tượng chưa có.",
        "review_alias_candidate": "Nếu best candidate đúng, thêm alias cụ thể; nếu sai, sửa extractor hoặc bổ sung danh mục.",
        "review_weak_alias": "Giữ alias yếu chỉ khi đã kiểm tra unique; nếu không thì thay bằng alias dài hơn.",
        "review_blocked_alias": "Alias quá ngắn hoặc giống mã chứng từ; giữ mã đối tượng nhưng thay alias bằng tên/số tài khoản/MST cụ thể hơn.",
        "improve_extractor": "Ưu tiên cải thiện bóc tách hint/counterparty/MST; không thêm alias từ dữ kiện rỗng hoặc quá chung.",
        "manual_object_review": "Review thủ công vì chưa có dữ kiện nhận diện đối tượng đủ mạnh.",
        "review_catalog_or_alias": "Kiểm tra danh mục trước; nếu có mã đúng thì thêm alias có kiểm soát.",
    }
    return actions.get(action_type, "")


def _catalog_for_flow(flow: Any) -> str:
    normalized = _review_norm(flow)
    if normalized == "BAO NO":
        return "payable"
    if normalized == "BAO CO":
        return "receivable"
    return ""


def _source_ref(record: dict[str, Any]) -> str:
    source_file = _clean_report_value(record.get("source_file"))
    source_row = _clean_report_value(record.get("source_row"))
    if not source_file and not source_row:
        return ""
    return f"{source_file}:{source_row}" if source_row else source_file


def _candidate_codes(value: Any) -> str:
    text = _clean_report_value(value)
    if not text:
        return ""
    codes = []
    for part in text.split(";"):
        code = part.strip().split("|", 1)[0].strip()
        if code and code not in codes:
            codes.append(code)
    return ", ".join(codes[:5])


def _clean_report_value(value: Any) -> str:
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    return str(value or "").strip()


def _hint_collision_records(processed: list[ProcessedTransaction]) -> list[dict[str, Any]]:
    return [
        _hint_collision_record(item)
        for item in processed
        if _has_hint_collision_signal(item)
    ]


def _has_hint_collision_signal(item: ProcessedTransaction) -> bool:
    candidates = item.matched_candidates
    if len(candidates) < 2:
        return False
    best = candidates[0]
    second = candidates[1]
    gap = float(best.score or 0) - float(second.score or 0)
    if item.status != "OK" and item.object_code == "ERROR":
        return True
    return gap < 8 and best.score >= 80


def _hint_collision_record(item: ProcessedTransaction) -> dict[str, Any]:
    best = item.matched_candidates[0]
    second = item.matched_candidates[1]
    gap = float(best.score or 0) - float(second.score or 0)
    normalized_error = _review_norm(item.error_note)
    risk_class = "low_gap_review"
    if item.status != "OK" and item.object_code == "ERROR":
        risk_class = "ambiguous_error" if "NHIEU MA DOI TUONG" in normalized_error else "object_error_with_candidates"
    hint = item.entities.counterparty_hint
    if not hint:
        risk_class = "weak_or_empty_hint"
    return {
        "risk_class": risk_class,
        "source_file": item.source_file,
        "source_sheet": item.source_sheet,
        "source_row": item.original_row_index,
        "bank": item.bank,
        "flow": item.flow,
        "counterparty_hint": hint,
        "counterparty_source": item.entities.counterparty_source,
        "object_match_source": item.object_match_source,
        "matched_object_code": item.object_code,
        "best_candidate_code": best.code,
        "best_candidate_score": best.score,
        "second_candidate_code": second.code,
        "second_candidate_score": second.score,
        "score_gap": round(gap, 4),
        "candidate_codes": ", ".join(candidate.code for candidate in item.matched_candidates[:5]),
        "object_ml_status": item.object_ml_result.status,
        "object_ml_best_code": item.object_ml_result.best_code,
        "object_ml_confidence": item.object_ml_result.confidence,
        "object_ml_gap": item.object_ml_result.gap,
        "object_ml_decision": item.object_ml_result.decision,
        "original_content": item.original_content,
        "error_note": item.error_note,
        "suggested_action": _hint_collision_suggested_action(risk_class),
        "transaction_uid": item.transaction_uid,
    }


def _hint_collision_suggested_action(risk_class: str) -> str:
    actions = {
        "ambiguous_error": "Chọn mã đúng nếu có dữ kiện; thêm alias mạnh hoặc giữ exception nếu hint thật sự đụng nhiều mã.",
        "object_error_with_candidates": "Review best candidate; nếu đúng thì thêm alias mạnh, nếu sai thì bổ sung dữ kiện hoặc danh mục.",
        "low_gap_review": "Kiểm tra lại vì top candidates gần điểm nhau; cân nhắc alias có MST/tên đầy đủ.",
        "weak_or_empty_hint": "Cải thiện extractor hoặc bổ sung dữ kiện nhận diện như MST/số tài khoản/tên đối tượng.",
    }
    return actions.get(risk_class, "")


def _rpa_record(
    item: ProcessedTransaction,
    run_id: str | None = None,
    reason_encoding: str = "",
) -> dict[str, Any]:
    record = {
        "Ngày CT": item.transaction_date,
        "Mã ĐT": item.object_code,
        "Lí do": _rpa_reason(item.reason, reason_encoding),
        "Người nộp tiền": _cash_recipient_name(item),
        "Người nhận tiền": _cash_recipient_name(item),
        "TK nợ": item.debit_account,
        "TK có": item.credit_account,
        "Thành tiền": item.amount,
        "Tỷ giá": item.exchange_rate or "",
        "Ngân hàng": item.bank,
        "transaction_uid": item.transaction_uid,
        "run_id": run_id or "",
        "rpa_status": item.rpa_status or STATUS_PENDING,
        "rpa_message": item.rpa_message,
        INPUT_STATUS_COLUMN: item.rpa_status or STATUS_PENDING,
        INPUT_MESSAGE_COLUMN: item.rpa_message or "",
        INPUT_UPDATED_AT_COLUMN: "",
        "source_file": item.source_file,
        "source_sheet": item.source_sheet,
        "source_row": item.original_row_index,
        "direction": item.bank_direction or item.flow,
    }
    if _is_tcvn3_reason_encoding(reason_encoding):
        record[RPA_REASON_UNICODE_COLUMN] = item.reason
    return record


def _rpa_reason(reason: str, encoding: str = "") -> str:
    if _is_tcvn3_reason_encoding(encoding):
        return unicode_to_tcvn3(reason)
    return reason


def _is_tcvn3_reason_encoding(encoding: str = "") -> bool:
    return str(encoding or "").strip().lower() == RPA_REASON_ENCODING_TCVN3


def _exception_record(item: ProcessedTransaction) -> dict[str, Any]:
    return {
        "Mã định danh": item.transaction_uid,
        "File gốc": item.source_file,
        "Sheet gốc": item.source_sheet,
        "Dòng gốc": item.original_row_index,
        "Ngân hàng": item.bank,
        "Luồng": item.flow,
        "Ngày CT": item.transaction_date,
        "Nội dung giao dịch gốc": item.original_content,
        "Người hưởng/Người chuyển": item.counterparty_raw,
        "Người nhận tiền": _cash_recipient_name(item),
        "Mã ĐT": item.object_code,
        "Tên ĐT suy luận": item.object_name,
        "TK nợ": item.debit_account,
        "TK có": item.credit_account,
        "Thành tiền": item.amount,
        "Ngoại tệ": item.foreign_currency,
        "Số tiền ngoại tệ": item.foreign_amount or "",
        "Tỷ giá": item.exchange_rate or "",
        "Use case dự đoán": item.use_case,
        "Trạng thái": item.status,
        "Trạng thái RPA": item.rpa_status,
        "Thông báo RPA": item.rpa_message,
        "transaction_uid": item.transaction_uid,
        "Ghi chú lỗi": item.error_note,
        "Độ tin cậy": item.confidence,
        "Nguồn match ĐT": item.object_match_source,
        "Counterparty hint": item.entities.counterparty_hint,
    }


def _manual_review_record(item: ProcessedTransaction) -> dict[str, Any]:
    return {
        "transaction_uid": item.transaction_uid,
        "source_file": item.source_file,
        "source_sheet": item.source_sheet,
        "source_row": item.original_row_index,
        "bank": item.bank,
        "bank_direction": item.bank_direction,
        "flow": item.flow,
        "transaction_date": item.transaction_date,
        "doc_no": item.doc_no,
        "description_raw": item.original_content,
        "description_normalized": item.normalized_content,
        "counterparty_raw": item.counterparty_raw,
        "cash_person_name": _cash_recipient_name(item),
        "amount": item.amount,
        "foreign_currency": item.foreign_currency,
        "foreign_amount": item.foreign_amount,
        "exchange_rate": item.exchange_rate,
        "rule_id": item.matched_rule,
        "voucher_type": flow_voucher_type(item.flow),
        "debit_account": item.debit_account,
        "credit_account": item.credit_account,
        "object_code": item.object_code,
        "confidence": item.confidence,
        "used_ml": "yes" if item.matched_rule == "ML" else "no",
        "review_reason": item.error_note,
        "is_duplicate": item.is_duplicate,
        "duplicate_of": item.duplicate_of,
    }


def _audit_record(item: ProcessedTransaction) -> dict[str, Any]:
    return {
        "transaction_uid": item.transaction_uid,
        "bank": item.bank,
        "source_file": item.source_file,
        "source_sheet": item.source_sheet,
        "source_row": item.original_row_index,
        "transaction_date": item.transaction_date,
        "description_raw": item.original_content,
        "description_normalized": item.normalized_content,
        "reference_no": item.doc_no,
        "counterparty_raw": item.counterparty_raw,
        "cash_person_name": _cash_recipient_name(item),
        "bank_direction": item.bank_direction,
        "amount": item.amount,
        "foreign_currency": item.foreign_currency,
        "foreign_amount": item.foreign_amount,
        "exchange_rate": item.exchange_rate,
        "rule_id": item.matched_rule,
        "flow": item.flow,
        "voucher_type": flow_voucher_type(item.flow),
        "debit_account": item.debit_account,
        "credit_account": item.credit_account,
        "object_code": item.object_code,
        "object_name": item.object_name,
        "confidence": item.confidence,
        "used_ml": "yes" if item.matched_rule == "ML" else "no",
        "status": item.status,
        "manual_review_reason": item.error_note if item.status != "OK" else "",
        "is_duplicate": item.is_duplicate,
        "duplicate_of": item.duplicate_of,
        "raw_data": item.raw_data,
    }


def _has_object_match_error(item: ProcessedTransaction) -> bool:
    error_note = str(item.error_note or "")
    normalized_error = _review_norm(error_note)
    return item.status != "OK" and ("MA DOI TUONG" in normalized_error or "MA DT" in normalized_error)


def _object_match_review_record(item: ProcessedTransaction) -> dict[str, Any]:
    reason_class = _object_match_reason_class(item)
    best = item.matched_candidates[0] if item.matched_candidates else None
    return {
        "reason_class": reason_class,
        "suggested_action": _object_match_suggested_action(reason_class),
        "error_note": item.error_note,
        "source_file": item.source_file,
        "source_sheet": item.source_sheet,
        "source_row": item.original_row_index,
        "bank": item.bank,
        "flow": item.flow,
        "bank_direction": item.bank_direction,
        "use_case": item.use_case,
        "original_content": item.original_content,
        "counterparty_raw": item.counterparty_raw,
        "counterparty_hint": item.entities.counterparty_hint,
        "counterparty_source": item.entities.counterparty_source,
        "tax_code": item.entities.tax_code,
        "object_match_source": item.object_match_source,
        "best_candidate_code": best.code if best else "",
        "best_candidate_name": best.name if best else "",
        "best_candidate_score": best.score if best else "",
        "best_candidate_source": best.source if best else "",
        "best_candidate_matched_on": best.matched_on if best else "",
        "object_ml_status": item.object_ml_result.status,
        "object_ml_best_code": item.object_ml_result.best_code,
        "object_ml_confidence": item.object_ml_result.confidence,
        "object_ml_gap": item.object_ml_result.gap,
        "object_ml_decision": item.object_ml_result.decision,
        "object_ml_note": item.object_ml_result.note,
        "top_candidates": _format_candidates(item.matched_candidates),
        "transaction_uid": item.transaction_uid,
    }


def _object_match_reason_class(item: ProcessedTransaction) -> str:
    error_note = _review_norm(item.error_note)
    hint = _review_norm(item.entities.counterparty_hint)
    tax_code = _review_norm(item.entities.tax_code)
    if "NHIEU MA DOI TUONG" in error_note or "NHIEU MA DT" in error_note:
        return "ambiguous"
    if tax_code and not any(candidate.source == "tax_code" for candidate in item.matched_candidates):
        return "tax_code_not_in_catalog"
    if hint in _WEAK_REVIEW_HINTS:
        return "weak_or_generic_hint"
    if not hint and not tax_code and not item.matched_candidates:
        return "no_extractable_object"
    if item.matched_candidates:
        best = item.matched_candidates[0]
        if best.score >= 70:
            return "missing_alias_or_low_score"
    return "not_in_catalog_or_missing_alias"


def _object_match_suggested_action(reason_class: str) -> str:
    actions = {
        "ambiguous": "Review mã đúng; thêm alias nếu một ứng viên rõ ràng, giữ exception nếu nhiều đối tượng thật.",
        "tax_code_not_in_catalog": "Kiểm tra danh mục Excel; bổ sung MST/mã đối tượng nếu thiếu.",
        "weak_or_generic_hint": "Cải thiện bóc tách hint từ nội dung/counterparty_raw; không match bằng cụm pháp lý chung.",
        "no_extractable_object": "Review thủ công nội dung sao kê vì chưa có dữ kiện nhận diện đối tượng.",
        "missing_alias_or_low_score": "Nếu best candidate đúng và mã có trong danh mục, bổ sung alias có kiểm soát.",
        "not_in_catalog_or_missing_alias": "Kiểm tra đối tượng trong danh mục; nếu có mã thì thêm alias, nếu không thì bổ sung danh mục.",
    }
    return actions.get(reason_class, "")


def _format_candidates(candidates: list[Any]) -> str:
    parts = []
    for candidate in candidates[:5]:
        ml_score = f"|ml={candidate.ml_score}" if getattr(candidate, "ml_score", 0) else ""
        parts.append(f"{candidate.code}|{candidate.score}|{candidate.source}|{candidate.matched_on}{ml_score}")
    return "; ".join(parts)


def _object_match_review_summary(detail_df: pd.DataFrame) -> pd.DataFrame:
    if detail_df.empty:
        return pd.DataFrame(columns=["group", "key", "count"])
    rows: list[dict[str, Any]] = []
    for group, column in (
        ("reason_class", "reason_class"),
        ("bank", "bank"),
        ("use_case", "use_case"),
        ("counterparty_hint", "counterparty_hint"),
    ):
        counts = detail_df[column].fillna("").astype(str).value_counts().head(50)
        for key, count in counts.items():
            rows.append({"group": group, "key": key, "count": int(count)})
    return pd.DataFrame(rows, columns=["group", "key", "count"])


def _review_norm(value: Any) -> str:
    from .normalizer import normalize_text

    return normalize_text(value)


_WEAK_REVIEW_HINTS = {
    "",
    "TRACH NHIEM HUU HAN",
    "CONG TY",
    "CONG TY TNHH",
    "CONG TY CO PHAN",
    "THANH TOAN",
    "DICH VU",
    "THUONG MAI",
    "VAN TAI",
}


def _task_records(
    flow_items: dict[str, list[ProcessedTransaction]],
    run_id: str | None,
) -> list[dict[str, Any]]:
    if not run_id:
        return []
    rows: list[dict[str, Any]] = []
    for flow in PAD_FLOW_ORDER:
        items = flow_items.get(flow, [])
        input_sheet = flow_sheet(flow)
        for excel_row, item in enumerate(items, start=2):
            rows.append(
                {
                    "run_id": run_id,
                    "task_id": f"{run_id}:{input_sheet}:{excel_row}",
                    "transaction_uid": item.transaction_uid,
                    "flow": item.flow,
                    "input_sheet": input_sheet,
                    "input_excel_row": excel_row,
                    "summary_status": item.rpa_status or STATUS_PENDING,
                    "source_file": item.source_file,
                    "source_sheet": item.source_sheet,
                    "source_row_index": item.original_row_index,
                    "sheet_row_count": len(items),
                    "execution_order": flow_execution_order(flow),
                    "voucher_type": flow_voucher_type(flow),
                    "has_error": "no",
                }
            )
    return rows


def _summary_records(
    processed: list[ProcessedTransaction],
    flow_items: dict[str, list[ProcessedTransaction]],
    run_stats: dict[str, Any] | None = None,
) -> list[dict[str, Any]]:
    run_stats = run_stats or {}
    rows: list[dict[str, Any]] = []

    def add(group: str, metric: str, value: Any) -> None:
        rows.append({"Nhóm": group, "Chỉ tiêu": metric, "Giá trị": value})

    add("Tổng quan", "Tổng số giao dịch đọc vào", len(processed))
    add("Tổng quan", "Số giao dịch đầu vào hợp lệ", run_stats.get("input_transaction_count", len(processed)))
    add("Tổng quan", "Số Báo nợ", sum(1 for item in processed if item.flow == "bao_no"))
    add("Tổng quan", "Số Báo có", sum(1 for item in processed if item.flow == "bao_co"))
    add("Tổng quan", "Số Phiếu thu tiền mặt", sum(1 for item in processed if item.flow == "thu_tien_mat"))
    add("Tổng quan", "Số Phiếu chi tiền mặt", sum(1 for item in processed if item.flow == "chi_tien_mat"))
    for flow in PAD_FLOW_ORDER:
        add("Tổng quan", f"Số dòng {flow_sheet(flow)}", len(flow_items.get(flow, [])))
    add("Tổng quan", "Số dòng đủ điều kiện nhập tự động", sum(len(items) for items in flow_items.values()))
    add("Tổng quan", "Số dòng chưa nhập", sum(1 for item in processed if item.rpa_status == "chua_nhap"))
    add("Tổng quan", "Số dòng hoàn thành", sum(1 for item in processed if item.rpa_status == "hoan_thanh"))
    add("Tổng quan", "Số dòng OK", sum(1 for item in processed if item.status == "OK"))
    add("Tổng quan", "Số giao dịch lỗi", sum(1 for item in processed if item.status != "OK"))
    add("Tổng quan", "Số giao dịch chờ kiểm tra", sum(1 for item in processed if item.status != "OK"))
    add("Tổng quan", "Số dòng bị bỏ qua vì là tiêu đề/số dư/tổng cộng", run_stats.get("skipped_non_transaction_rows", 0))
    add("Tổng quan", "Số giao dịch trùng", run_stats.get("duplicate_count", sum(1 for item in processed if item.is_duplicate)))

    total_money_in = sum(item.amount for item in processed if item.bank_direction == "money_in")
    total_money_out = sum(item.amount for item in processed if item.bank_direction == "money_out")
    add("Tổng tiền", "Tổng tiền vào", total_money_in)
    add("Tổng tiền", "Tổng tiền ra", total_money_out)
    for flow in PAD_FLOW_ORDER:
        add("Tổng tiền từng luồng", flow_sheet(flow), sum(item.amount for item in processed if item.flow == flow and item.status == "OK"))

    for bank in sorted({item.bank for item in processed if item.bank}):
        add(
            "Tổng tiền OK theo ngân hàng",
            bank,
            sum(item.amount for item in processed if item.bank == bank and item.status == "OK"),
        )
        add(
            "Tổng tiền EXCEPTION theo ngân hàng",
            bank,
            sum(item.amount for item in processed if item.bank == bank and item.status != "OK"),
        )

    for use_case in sorted({item.use_case for item in processed if item.use_case}):
        add("Số dòng theo use case", use_case, sum(1 for item in processed if item.use_case == use_case))

    for status in sorted({item.status for item in processed if item.status}):
        add("Số dòng theo trạng thái", status, sum(1 for item in processed if item.status == status))

    for flow in sorted({item.flow for item in processed if item.flow}):
        add("Tổng tiền theo luồng", flow, sum(item.amount for item in processed if item.flow == flow))

    out_ok = sum(item.amount for item in processed if item.bank_direction == "money_out" and item.status == "OK")
    out_error = sum(item.amount for item in processed if item.bank_direction == "money_out" and item.status != "OK")
    in_ok = sum(item.amount for item in processed if item.bank_direction == "money_in" and item.status == "OK")
    in_error = sum(item.amount for item in processed if item.bank_direction == "money_in" and item.status != "OK")
    add("Cân đối", "Chênh lệch tiền ra", round(total_money_out - out_ok - out_error, 2))
    add("Cân đối", "Chênh lệch tiền vào", round(total_money_in - in_ok - in_error, 2))
    return rows


def _tracking_record(item: ProcessedTransaction) -> dict[str, Any]:
    return {
        "transaction_uid": item.transaction_uid,
        "bank_code": item.bank,
        "source_file": item.source_file,
        "source_sheet": item.source_sheet,
        "source_row": item.original_row_index,
        "direction": item.flow,
        "rpa_status": item.rpa_status or STATUS_PENDING,
        "rpa_message": item.rpa_message,
        "created_at": "",
        "updated_at": "",
        "completed_at": "",
        "original_row_index": item.original_row_index,
        "bank": item.bank,
        "flow": item.flow,
        "bank_direction": item.bank_direction,
        "transaction_date": _serialize_date(item.transaction_date),
        "doc_no": item.doc_no,
        "original_content": item.original_content,
        "normalized_content": item.normalized_content,
        "counterparty_raw": item.counterparty_raw,
        "cash_person_name": _cash_recipient_name(item),
        "normalized_counterparty": item.normalized_counterparty,
        "entities": asdict(item.entities),
        "matched_object_code": item.object_code,
        "matched_object_name": item.object_name,
        "matched_candidates": [asdict(candidate) for candidate in item.matched_candidates],
        "object_match_source": item.object_match_source,
        "reason": item.reason,
        "debit_account": item.debit_account,
        "credit_account": item.credit_account,
        "amount": item.amount,
        "foreign_currency": item.foreign_currency,
        "foreign_amount": item.foreign_amount,
        "exchange_rate": item.exchange_rate,
        "use_case": item.use_case,
        "matched_rule": item.matched_rule,
        "ml_result": asdict(item.ml_result),
        "object_ml_result": asdict(item.object_ml_result),
        "verification_result": asdict(item.verification_result),
        "processing_status": item.status,
        "error_note": item.error_note,
        "confidence": item.confidence,
        "raw_data": item.raw_data,
        "is_duplicate": item.is_duplicate,
        "duplicate_of": item.duplicate_of,
    }


def _tracking_record_from_summary(summary_record: dict[str, Any]) -> dict[str, Any]:
    record = {
        "transaction_uid": summary_record.get("transaction_uid", ""),
        "bank_code": summary_record.get("bank_code") or summary_record.get("bank", ""),
        "source_file": summary_record.get("source_file", ""),
        "source_sheet": summary_record.get("source_sheet", ""),
        "source_row": summary_record.get("source_row") or summary_record.get("source_row_index", ""),
        "direction": summary_record.get("direction") or summary_record.get("flow", ""),
        "rpa_status": summary_record.get("rpa_status") or summary_record.get("status", ""),
        "rpa_message": summary_record.get("rpa_message", ""),
        "created_at": summary_record.get("created_at", ""),
        "updated_at": summary_record.get("updated_at", ""),
        "completed_at": summary_record.get("completed_at", ""),
        "bank": summary_record.get("bank") or summary_record.get("bank_code", ""),
        "flow": summary_record.get("flow") or summary_record.get("direction", ""),
        "source_row_index": summary_record.get("source_row_index") or summary_record.get("source_row", ""),
        "transaction_date": _serialize_cell(summary_record.get("transaction_date")),
        "doc_no": summary_record.get("doc_no", ""),
        "original_content": summary_record.get("original_content", ""),
        "counterparty_raw": summary_record.get("counterparty_raw", ""),
        "cash_person_name": summary_record.get("cash_person_name", ""),
        "matched_object_code": summary_record.get("object_code", ""),
        "matched_object_name": summary_record.get("object_name", ""),
        "reason": summary_record.get("reason", ""),
        "debit_account": summary_record.get("debit_account", ""),
        "credit_account": summary_record.get("credit_account", ""),
        "amount": summary_record.get("amount", ""),
    }
    return {key: _serialize_cell(value) for key, value in record.items()}


def _apply_summary_tracking_fields(record: dict[str, Any], summary_record: dict[str, Any]) -> None:
    mapping = {
        "bank_code": summary_record.get("bank_code") or summary_record.get("bank", ""),
        "source_row": summary_record.get("source_row") or summary_record.get("source_row_index", ""),
        "direction": summary_record.get("direction") or summary_record.get("flow", ""),
        "rpa_status": summary_record.get("rpa_status") or summary_record.get("status", ""),
        "rpa_message": summary_record.get("rpa_message", ""),
        "created_at": summary_record.get("created_at", ""),
        "updated_at": summary_record.get("updated_at", ""),
        "completed_at": summary_record.get("completed_at", ""),
        "rpa_started_at": summary_record.get("rpa_started_at", ""),
        "rpa_finished_at": summary_record.get("rpa_finished_at", ""),
        "voucher_no": summary_record.get("voucher_no", ""),
        "last_run_id": summary_record.get("last_run_id", ""),
        "last_attempt_result": summary_record.get("last_attempt_result", ""),
        "status": summary_record.get("rpa_status") or summary_record.get("status", ""),
    }
    for key, value in mapping.items():
        record[key] = _serialize_cell(value)


def _cash_recipient_name(item: ProcessedTransaction) -> str:
    if item.flow not in {FLOW_THU_TIEN_MAT, FLOW_CHI_TIEN_MAT}:
        return ""
    return str(getattr(item.entities, "cash_person_name", "") or "").strip()


def _rpa_columns_for_flow(flow: str, reason_encoding: str = "") -> list[str]:
    if flow == FLOW_THU_TIEN_MAT:
        columns = list(RPA_THU_TIEN_MAT_COLUMNS)
    else:
        columns = list(RPA_BUSINESS_COLUMNS)
    if _is_tcvn3_reason_encoding(reason_encoding):
        reason_index = columns.index("Lí do") + 1
        columns.insert(reason_index, RPA_REASON_UNICODE_COLUMN)
    return columns


def _serialize_date(value: date | None) -> str:
    return value.isoformat() if value else ""


def _serialize_cell(value: Any) -> Any:
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    if isinstance(value, date):
        return value.isoformat()
    return value


def _format_sheet(ws) -> None:
    for cell in ws[1]:
        cell.font = Font(bold=True)
    header_names = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
    for name, idx in header_names.items():
        if name and "Ngày" in str(name):
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=idx).number_format = "DD/MM/YYYY"
        if name in {"Thành tiền", "Giá trị", "Tỷ giá"}:
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=idx).number_format = "#,##0"

    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[letter]:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 60)
