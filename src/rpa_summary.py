from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .flows import FLOW_BAO_CO, FLOW_BAO_NO, FLOW_CHI_TIEN_MAT, FLOW_THU_TIEN_MAT
from .models import ProcessedTransaction
from .rpa_tracking import (
    ELIGIBLE_RPA_STATUSES,
    ATTEMPT_SUCCESS,
    STATUS_DONE,
    STATUS_PENDING,
    abort_run_records,
    apply_status_update,
    finalize_run_records,
    normalize_status,
    reset_all_records,
    validate_status,
)
from .transaction_identity import assign_transaction_uids


SUMMARY_SHEET_NAME = "RPA_SUMMARY"

SUMMARY_COLUMNS = [
    "transaction_uid",
    "rpa_status",
    "source_file",
    "voucher_no",
    "rpa_message",
    "transaction_date",
    "bank",
    "flow",
    "amount",
    "object_code",
    "debit_account",
    "credit_account",
    "reason",
    "original_content",
    "source_sheet",
    "source_row",
    "doc_no",
    "counterparty_raw",
    "object_name",
    "use_case",
    "processing_status",
    "confidence",
    "object_match_source",
    "bank_code",
    "direction",
    "created_at",
    "updated_at",
    "completed_at",
    "last_run_id",
    "rpa_started_at",
    "rpa_finished_at",
    "last_attempt_result",
    # Backward-compatible aliases used by older output and tests.
    "status",
    "source_row_index",
]

STATUS_GUIDE_ROWS = [
    {"field": "rpa_status", "editable": "YES", "note": "Chỉ dùng chua_nhap hoặc hoan_thanh."},
    {"field": "voucher_no", "editable": "YES", "note": "Số chứng từ VACOM nếu đã nhập xong."},
    {"field": "rpa_message", "editable": "YES", "note": "Ghi chú lỗi hoặc lý do cần nhập lại."},
    {"field": "transaction_uid", "editable": "NO", "note": "Mã định danh dòng, không sửa."},
    {"field": "Các cột còn lại", "editable": "NO", "note": "Dùng để đối chiếu sao kê và chứng từ."},
]


@dataclass
class RpaRunState:
    run_id: str
    summary_df: pd.DataFrame
    rpa_items: list[ProcessedTransaction]
    stats: dict[str, int] = field(default_factory=dict)


def make_run_id() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def prepare_rpa_run(
    processed: list[ProcessedTransaction],
    summary_path: str | Path,
    tracking_path: str | Path | None = None,
    run_id: str | None = None,
    logger: Any | None = None,
) -> RpaRunState:
    run_id = run_id or make_run_id()
    if any(not item.transaction_uid for item in processed):
        assign_transaction_uids(processed)

    existing_df = load_summary(summary_path)
    existing_by_uid = {
        str(row.get("transaction_uid", "")).strip(): _clean_record(row)
        for row in existing_df.to_dict("records")
        if str(row.get("transaction_uid", "")).strip()
    }

    now = _now()
    records_by_uid: dict[str, dict[str, Any]] = {}
    rpa_items: list[ProcessedTransaction] = []
    seen: set[str] = set()
    stats = {
        "total_processed": len(processed),
        "new_count": 0,
        "auto_process_count": 0,
        "pending_count": 0,
        "in_progress_count": 0,
        "error_count": 0,
        "skipped_count": 0,
        "skipped_completed_count": 0,
        "waiting_count": 0,
        "retry_error_count": 0,
        "review_count": 0,
        "exception_count": sum(1 for item in processed if item.status != "OK"),
        "bao_no_output_count": 0,
        "bao_co_output_count": 0,
        "thu_tien_mat_output_count": 0,
        "chi_tien_mat_output_count": 0,
    }

    for item in processed:
        uid = item.transaction_uid
        if not uid:
            continue
        previous = existing_by_uid.get(uid)
        if previous is None:
            stats["new_count"] += 1

        record = _merge_record(previous, item, run_id, now)
        records_by_uid[uid] = record
        seen.add(uid)

        item.rpa_status = str(record.get("rpa_status", "")).strip()
        item.rpa_message = str(record.get("rpa_message", "") or "")

        if item.rpa_status == STATUS_PENDING:
            stats["pending_count"] += 1
        if item.rpa_status == STATUS_DONE:
            stats["skipped_completed_count"] += 1
        if item.status == "OK" and item.rpa_status in ELIGIBLE_RPA_STATUSES:
            rpa_items.append(item)
            stats["auto_process_count"] += 1
            stats["waiting_count"] += 1
            if item.flow == FLOW_BAO_NO:
                stats["bao_no_output_count"] += 1
            elif item.flow == FLOW_BAO_CO:
                stats["bao_co_output_count"] += 1
            elif item.flow == FLOW_THU_TIEN_MAT:
                stats["thu_tien_mat_output_count"] += 1
            elif item.flow == FLOW_CHI_TIEN_MAT:
                stats["chi_tien_mat_output_count"] += 1

    for row in existing_df.to_dict("records"):
        uid = str(row.get("transaction_uid", "")).strip()
        if uid and uid not in seen:
            records_by_uid[uid] = _ensure_summary_record(_clean_record(row))

    summary_df = pd.DataFrame(list(records_by_uid.values()), columns=SUMMARY_COLUMNS)
    summary_df = _ensure_columns(summary_df)
    return RpaRunState(run_id=run_id, summary_df=summary_df, rpa_items=rpa_items, stats=stats)


def load_summary(path: str | Path) -> pd.DataFrame:
    path = Path(path)
    if not path.exists():
        return pd.DataFrame(columns=SUMMARY_COLUMNS)
    try:
        df = pd.read_excel(path, sheet_name=SUMMARY_SHEET_NAME, dtype=object)
    except ValueError:
        df = pd.read_excel(path, sheet_name=0, dtype=object)
    return _ensure_columns(df)


def write_summary(summary_df: pd.DataFrame, path: str | Path) -> None:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    df = _ensure_columns(summary_df)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=SUMMARY_SHEET_NAME, index=False)
        status_df = _status_counts(df)
        status_df.to_excel(writer, sheet_name="STATUS_COUNTS", index=False)
        guide_df = pd.DataFrame(STATUS_GUIDE_ROWS)
        guide_df.to_excel(writer, sheet_name="STATUS_GUIDE", index=False)
        _apply_summary_status_controls(writer.book[SUMMARY_SHEET_NAME])
        for sheet_name in (SUMMARY_SHEET_NAME, "STATUS_COUNTS", "STATUS_GUIDE"):
            _format_sheet(writer.book[sheet_name])


def update_rpa_status(
    summary_path: str | Path,
    transaction_uid: str,
    status: str,
    run_id: str = "",
    message: str = "",
    voucher_no: str = "",
) -> dict[str, Any]:
    status = validate_status(status)
    df = load_summary(summary_path)
    uid = str(transaction_uid).strip()
    matches = df.index[df["transaction_uid"].astype(str) == uid].tolist()
    if not matches:
        raise KeyError(f"transaction_uid not found in RPA summary: {uid}")

    row_idx = matches[0]
    record = _clean_record(df.loc[row_idx].to_dict())
    updated = apply_status_update(record, status, message=message, voucher_no=voucher_no, run_id=run_id)
    for column in SUMMARY_COLUMNS:
        df.at[row_idx, column] = updated.get(column, "")
    write_summary(df, summary_path)
    return _ensure_summary_record(updated)


def mark_rpa_started(summary_path: str | Path, transaction_uid: str, run_id: str) -> None:
    update_rpa_status(summary_path, transaction_uid, STATUS_PENDING, run_id=run_id)


def mark_rpa_done(
    summary_path: str | Path,
    transaction_uid: str,
    run_id: str,
    voucher_no: str = "",
    message: str = "",
) -> None:
    update_rpa_status(summary_path, transaction_uid, STATUS_DONE, run_id=run_id, message=message, voucher_no=voucher_no)


def mark_rpa_error(summary_path: str | Path, transaction_uid: str, run_id: str, message: str) -> None:
    update_rpa_status(summary_path, transaction_uid, STATUS_PENDING, run_id=run_id, message=message)


def finalize_rpa_run(summary_path: str | Path, run_id: str) -> pd.DataFrame:
    df = load_summary(summary_path)
    records = [_ensure_summary_record(record) for record in df.to_dict("records")]
    finalized = [_ensure_summary_record(record) for record in finalize_run_records(records, run_id)]
    result = pd.DataFrame(finalized, columns=SUMMARY_COLUMNS)
    write_summary(result, summary_path)
    return result


def abort_rpa_run(summary_path: str | Path, run_id: str, message: str = "") -> pd.DataFrame:
    df = load_summary(summary_path)
    records = [_ensure_summary_record(record) for record in df.to_dict("records")]
    aborted = [_ensure_summary_record(record) for record in abort_run_records(records, run_id, message=message)]
    result = pd.DataFrame(aborted, columns=SUMMARY_COLUMNS)
    write_summary(result, summary_path)
    return result


def reset_all_rpa_status(summary_path: str | Path, message: str = "") -> pd.DataFrame:
    df = load_summary(summary_path)
    records = [_ensure_summary_record(record) for record in df.to_dict("records")]
    reset_records = [_ensure_summary_record(record) for record in reset_all_records(records, message=message)]
    result = pd.DataFrame(reset_records, columns=SUMMARY_COLUMNS)
    write_summary(result, summary_path)
    return result


def _merge_record(
    previous: dict[str, Any] | None,
    item: ProcessedTransaction,
    run_id: str,
    now: str,
) -> dict[str, Any]:
    current = _record_from_item(item, run_id, now)
    if not previous:
        current["rpa_status"] = STATUS_PENDING
        current["status"] = current["rpa_status"]
        current["rpa_message"] = "" if item.status == "OK" else item.error_note
        return current

    previous = _ensure_summary_record(previous)
    previous_status = normalize_status(previous.get("rpa_status") or previous.get("status"), default=STATUS_PENDING)

    if previous_status == STATUS_DONE:
        return previous

    record = {**current, "created_at": previous.get("created_at") or now}
    _preserve_rpa_fields(record, previous)

    if item.status != "OK":
        record["rpa_status"] = STATUS_PENDING
        record["status"] = STATUS_PENDING
        record["rpa_message"] = item.error_note
        return record

    record["rpa_status"] = previous_status if previous_status == STATUS_DONE else STATUS_PENDING
    record["status"] = record["rpa_status"]
    if record["rpa_status"] == STATUS_PENDING:
        record["rpa_message"] = ""
    return record


def _record_from_item(item: ProcessedTransaction, run_id: str, now: str) -> dict[str, Any]:
    source_row = item.original_row_index
    return {
        "transaction_uid": item.transaction_uid,
        "bank_code": item.bank,
        "source_file": item.source_file,
        "source_sheet": item.source_sheet,
        "source_row": source_row,
        "direction": item.flow,
        "rpa_status": "",
        "rpa_message": "",
        "created_at": now,
        "updated_at": now,
        "completed_at": "",
        "transaction_date": item.transaction_date,
        "doc_no": item.doc_no,
        "original_content": item.original_content,
        "counterparty_raw": item.counterparty_raw,
        "amount": item.amount,
        "object_code": item.object_code,
        "object_name": item.object_name,
        "debit_account": item.debit_account,
        "credit_account": item.credit_account,
        "reason": item.reason,
        "use_case": item.use_case,
        "processing_status": item.status,
        "confidence": item.confidence,
        "object_match_source": item.object_match_source,
        "last_run_id": run_id,
        "rpa_started_at": "",
        "rpa_finished_at": "",
        "voucher_no": "",
        "last_attempt_result": "",
        "status": "",
        "bank": item.bank,
        "flow": item.flow,
        "source_row_index": source_row,
    }


def _preserve_rpa_fields(record: dict[str, Any], previous: dict[str, Any]) -> None:
    for field in ("rpa_started_at", "rpa_finished_at", "rpa_message", "voucher_no", "completed_at", "last_attempt_result"):
        record[field] = previous.get(field, "")


def _ensure_columns(df: pd.DataFrame) -> pd.DataFrame:
    result = df.copy()
    for column in SUMMARY_COLUMNS:
        if column not in result.columns:
            result[column] = ""

    result = result[SUMMARY_COLUMNS].astype(object)
    for idx, row in result.iterrows():
        record = _ensure_summary_record(_clean_record(row.to_dict()))
        for column in SUMMARY_COLUMNS:
            result.at[idx, column] = record.get(column, "")
    return result


def _ensure_summary_record(row: dict[str, Any]) -> dict[str, Any]:
    record = {column: _clean_cell(row.get(column, "")) for column in SUMMARY_COLUMNS}
    record["bank_code"] = record.get("bank_code") or row.get("bank") or ""
    record["bank"] = record.get("bank") or record["bank_code"]
    record["direction"] = record.get("direction") or row.get("flow") or ""
    record["flow"] = record.get("flow") or record["direction"]
    record["source_row"] = record.get("source_row") or row.get("source_row_index") or row.get("original_row_index") or ""
    record["source_row_index"] = record.get("source_row_index") or record["source_row"]
    rpa_status = normalize_status(record.get("rpa_status") or row.get("status"), default=STATUS_PENDING)
    if rpa_status != STATUS_DONE and str(record.get("last_attempt_result") or "") == ATTEMPT_SUCCESS:
        rpa_status = STATUS_DONE
        record["last_attempt_result"] = ""
    record["rpa_status"] = rpa_status
    record["status"] = rpa_status
    if rpa_status == STATUS_DONE and not record.get("completed_at"):
        record["completed_at"] = record.get("rpa_finished_at") or record.get("updated_at") or _now()
    return record


def _clean_record(row: dict[str, Any]) -> dict[str, Any]:
    return {str(key): _clean_cell(value) for key, value in row.items()}


def _clean_cell(value: Any) -> Any:
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    return value


def _status_counts(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=["rpa_status", "count"])
    counts = df["rpa_status"].fillna("").astype(str).value_counts().reset_index()
    counts.columns = ["rpa_status", "count"]
    return counts


def _now() -> str:
    return datetime.now().isoformat(timespec="seconds")


def _format_sheet(ws) -> None:
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    if ws.max_row and ws.max_column:
        ws.auto_filter.ref = ws.dimensions
    header_names = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
    for name, idx in header_names.items():
        if name and ("date" in str(name).lower() or str(name) in {"created_at", "updated_at", "completed_at"}):
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=idx).number_format = "DD/MM/YYYY"
        if name in {"amount", "count"}:
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=idx).number_format = "#,##0"

    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[letter]:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 60)


def _apply_summary_status_controls(ws) -> None:
    header_names = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
    status_idx = header_names.get("rpa_status")
    if not status_idx:
        return
    letter = get_column_letter(status_idx)
    validation = DataValidation(type="list", formula1='"chua_nhap,hoan_thanh"', allow_blank=False)
    validation.error = "Chỉ chọn chua_nhap hoặc hoan_thanh"
    validation.errorTitle = "Trạng thái không hợp lệ"
    validation.prompt = "Chọn chua_nhap hoặc hoan_thanh"
    validation.promptTitle = "rpa_status"
    ws.add_data_validation(validation)
    validation.add(f"{letter}2:{letter}1048576")

    pending_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    done_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    status_range = f"{letter}2:{letter}{max(ws.max_row, 2)}"
    ws.conditional_formatting.add(status_range, CellIsRule(operator="equal", formula=['"chua_nhap"'], fill=pending_fill))
    ws.conditional_formatting.add(status_range, CellIsRule(operator="equal", formula=['"hoan_thanh"'], fill=done_fill))
