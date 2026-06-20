from __future__ import annotations

from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .rpa_tracking import STATUS_DONE, validate_status


INPUT_STATUS_COLUMN = "Trạng thái RPA"
INPUT_MESSAGE_COLUMN = "Thông báo RPA"
INPUT_UPDATED_AT_COLUMN = "Thời gian nhập RPA"
TRANSACTION_UID_COLUMN = "transaction_uid"
RPA_INPUT_SHEETS = [
    "BAO_NO_INPUT",
    "BAO_CO_INPUT",
    "CHI_TIEN_MAT_INPUT",
    "THU_TIEN_MAT_INPUT",
]
INPUT_STATUS_COLUMNS = [
    INPUT_STATUS_COLUMN,
    INPUT_MESSAGE_COLUMN,
    INPUT_UPDATED_AT_COLUMN,
]


class RpaInputStatusError(RuntimeError):
    """Raised when an input workbook cannot be synchronized safely."""


def update_input_file_status(
    input_path: str | Path,
    transaction_uid: str,
    status: str,
    message: str = "",
    voucher_no: str = "",
    run_id: str = "",
) -> dict[str, Any]:
    input_path = Path(input_path)
    if not input_path.exists():
        raise FileNotFoundError(f"RPA input file not found: {input_path}")

    uid = str(transaction_uid or "").strip()
    if not uid:
        raise RpaInputStatusError("transaction_uid is required")

    normalized_status = validate_status(status)
    updated_at = _now()
    workbook = load_workbook(input_path)
    updated_rows: list[dict[str, Any]] = []

    for ws in iter_rpa_input_sheets(workbook):
        header_columns = find_header_columns(ws)
        uid_column = header_columns.get(TRANSACTION_UID_COLUMN)
        if not uid_column:
            continue

        header_columns = ensure_input_status_columns(ws)
        uid_column = header_columns[TRANSACTION_UID_COLUMN]
        for row_index in range(2, ws.max_row + 1):
            if _clean_uid(ws.cell(row=row_index, column=uid_column).value) != uid:
                continue

            ws.cell(row=row_index, column=header_columns[INPUT_STATUS_COLUMN], value=normalized_status)
            ws.cell(row=row_index, column=header_columns[INPUT_MESSAGE_COLUMN], value=message or "")
            time_value = updated_at if normalized_status == STATUS_DONE else ""
            ws.cell(row=row_index, column=header_columns[INPUT_UPDATED_AT_COLUMN], value=time_value)
            updated_rows.append({"sheet": ws.title, "row": row_index})

    if not updated_rows:
        raise RpaInputStatusError(f"transaction_uid not found in input file: {uid}")

    workbook.save(input_path)
    return {
        "input_path": input_path,
        "transaction_uid": uid,
        "status": normalized_status,
        "updated_rows": updated_rows,
        "voucher_no": voucher_no,
        "run_id": run_id,
    }


def ensure_input_status_columns(ws) -> dict[str, int]:
    header_columns = find_header_columns(ws)
    last_column = _last_header_column(ws, header_columns)
    for column_name in INPUT_STATUS_COLUMNS:
        if column_name in header_columns:
            continue
        last_column += 1
        ws.cell(row=1, column=last_column, value=column_name)
        _copy_header_style(ws, last_column - 1, last_column)
        header_columns[column_name] = last_column
    return header_columns


def find_header_columns(ws) -> dict[str, int]:
    columns: dict[str, int] = {}
    for idx, cell in enumerate(ws[1], start=1):
        name = str(cell.value or "").strip()
        if name and name not in columns:
            columns[name] = idx
    return columns


def iter_rpa_input_sheets(workbook):
    for sheet_name in RPA_INPUT_SHEETS:
        if sheet_name in workbook.sheetnames:
            yield workbook[sheet_name]


def _last_header_column(ws, header_columns: dict[str, int]) -> int:
    if not header_columns:
        return 0
    return max(max(header_columns.values()), ws.max_column)


def _clean_uid(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _copy_header_style(ws, source_column: int, target_column: int) -> None:
    if source_column < 1:
        return
    source = ws.cell(row=1, column=source_column)
    target = ws.cell(row=1, column=target_column)
    if source.has_style:
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.protection = copy(source.protection)
        target.number_format = source.number_format


def _now() -> str:
    return datetime.now().isoformat(timespec="seconds")
